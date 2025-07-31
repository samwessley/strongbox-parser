#!/usr/bin/env python3

import pandas as pd
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
import calendar
import openpyxl
import warnings
warnings.filterwarnings('ignore')

def debug_tb_data(filename):
    """Debug the TB-DATA sheet contents"""
    print("\n🔍 DEBUGGING TB-DATA SHEET")
    print("=" * 50)
    
    try:
        # Read TB-DATA sheet directly
        tb_data = pd.read_excel(filename, sheet_name='TB-DATA')
        print(f"📊 TB-DATA has {len(tb_data)} rows and {len(tb_data.columns)} columns")
        print(f"📋 Columns: {list(tb_data.columns)}")
        
        # Show first few rows
        print("\n📋 First 5 rows of TB-DATA:")
        print(tb_data.head().to_string())
        
        # Check for non-zero balances
        start_bal_col = None
        end_bal_col = None
        
        # Find balance columns
        for col in tb_data.columns:
            if 'start' in str(col).lower() and 'balance' in str(col).lower():
                start_bal_col = col
            elif 'end' in str(col).lower() and 'balance' in str(col).lower():
                end_bal_col = col
        
        if start_bal_col and end_bal_col:
            print(f"\n💰 Balance columns found:")
            print(f"  • Starting balance column: {start_bal_col}")
            print(f"  • Ending balance column: {end_bal_col}")
            
            # Check balance values
            start_sum = tb_data[start_bal_col].sum()
            end_sum = tb_data[end_bal_col].sum()
            print(f"\n💰 Total balances in TB-DATA:")
            print(f"  • Sum of starting balances: {start_sum:,.2f}")
            print(f"  • Sum of ending balances: {end_sum:,.2f}")
            
            # Show non-zero balances
            non_zero = tb_data[(tb_data[start_bal_col] != 0) | (tb_data[end_bal_col] != 0)]
            print(f"\n💰 Found {len(non_zero)} accounts with non-zero balances")
            if len(non_zero) > 0:
                print("\n📋 Sample of accounts with balances:")
                for _, row in non_zero.head().iterrows():
                    print(f"  • Account {row['Account Id']}: Start={row[start_bal_col]:,.2f}, End={row[end_bal_col]:,.2f}")
        
    except Exception as e:
        print(f"❌ Error reading TB-DATA: {str(e)}")

def debug_transactions(filename):
    """Debug the transaction sheets"""
    print("\n🔍 DEBUGGING TRANSACTION SHEETS")
    print("=" * 50)
    
    try:
        # Get list of transaction sheets
        wb = openpyxl.load_workbook(filename, read_only=True)
        txn_sheets = [s for s in wb.sheetnames if s.startswith('TXN-FY')]
        print(f"📋 Found {len(txn_sheets)} transaction sheets: {txn_sheets}")
        wb.close()
        
        # Read each transaction sheet
        for sheet_name in txn_sheets:
            print(f"\n📋 Reading {sheet_name}:")
            try:
                df = pd.read_excel(filename, sheet_name=sheet_name)
                print(f"  • Sheet has {len(df)} rows and {len(df.columns)} columns")
                print(f"  • Columns: {list(df.columns)}")
                
                # Check transaction dates
                if 'Fiscal Month' in df.columns:
                    dates = pd.to_datetime(df['Fiscal Month'], errors='coerce')
                    print("\n  📅 Fiscal Month dates found:")
                    for date in dates.unique():
                        count = len(df[df['Fiscal Month'] == date])
                        print(f"    • {date}: {count} transactions")
                
                # Show sample transactions
                print("\n  💰 Sample transactions:")
                for idx, row in df.iterrows():
                    debit = float(row.get('Debit', 0) or 0)
                    credit = float(row.get('Credit', 0) or 0)
                    print(f"    • Row {idx+1}: Account={row.get('Account Id', 'N/A')}, Debit={debit:,.2f}, Credit={credit:,.2f}")
                    if idx >= 4:  # Show first 5 rows
                        break
                
            except Exception as e:
                print(f"  ❌ Error reading sheet: {str(e)}")
    
    except Exception as e:
        print(f"❌ Error accessing transaction sheets: {str(e)}")

def debug_output_file(filename):
    """Debug the generated output file"""
    print("\n🔍 DEBUGGING OUTPUT FILE")
    print("=" * 50)
    
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        print(f"📋 Sheets in output: {wb.sheetnames}")
        
        # Check Comparative Trial Balances
        if 'Comparative Trial Balances' in wb.sheetnames:
            tb_sheet = wb['Comparative Trial Balances']
            print(f"\n📊 COMPARATIVE TRIAL BALANCES:")
            print(f"  • Sheet has {tb_sheet.max_row} rows and {tb_sheet.max_column} columns")
            
            # Check balances
            non_zero_count = 0
            total_begin = 0
            total_end = 0
            
            for row in range(3, tb_sheet.max_row + 1):  # Skip headers
                begin_val = tb_sheet.cell(row=row, column=3).value or 0
                end_val = tb_sheet.cell(row=row, column=4).value or 0
                
                total_begin += float(begin_val)
                total_end += float(end_val)
                
                if begin_val != 0 or end_val != 0:
                    non_zero_count += 1
                    if non_zero_count <= 5:  # Show first 5 non-zero rows
                        account = tb_sheet.cell(row=row, column=1).value
                        print(f"    • Row {row}: Account={account}, Begin={begin_val:,.2f}, End={end_val:,.2f}")
            
            print(f"\n  💰 Balance Summary:")
            print(f"    • Non-zero balance rows: {non_zero_count}")
            print(f"    • Total Beginning Balance: {total_begin:,.2f}")
            print(f"    • Total Ending Balance: {total_end:,.2f}")
        
        # Check Journal Entries
        je_sheets = [s for s in wb.sheetnames if s.startswith('Journal Entries & Lines')]
        print(f"\n📝 JOURNAL ENTRIES SHEETS:")
        
        for sheet_name in je_sheets:
            sheet = wb[sheet_name]
            print(f"\n  📋 {sheet_name}:")
            print(f"    • Sheet has {sheet.max_row} rows")
            
            # Show transactions
            total_debit = 0
            total_credit = 0
            
            for row in range(3, sheet.max_row + 1):  # Skip headers
                debit = sheet.cell(row=row, column=8).value or 0  # Debit Amount column
                credit = sheet.cell(row=row, column=9).value or 0  # Credit Amount column
                
                total_debit += float(debit)
                total_credit += float(credit)
                
                if row <= 7:  # Show first 5 transactions
                    account = sheet.cell(row=row, column=5).value  # Account ID column
                    print(f"    • Row {row}: Account={account}, Debit={debit:,.2f}, Credit={credit:,.2f}")
            
            print(f"\n    💰 Totals:")
            print(f"      • Total Debits: {total_debit:,.2f}")
            print(f"      • Total Credits: {total_credit:,.2f}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading output file: {str(e)}")

def main():
    input_file = "STRONGBOX FULL ANALYSIS - ACCRUAL BASIS - ATS ENGINEERS, INSPECTORS, & SURVEYORS V1.XLSX"
    output_file = "Processed_Strongbox_20230201_20250716.xlsx"
    
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        return
        
    print("🔍 Starting debug process...")
    debug_tb_data(input_file)
    debug_transactions(input_file)
    
    if os.path.exists(output_file):
        debug_output_file(output_file)
    else:
        print(f"\n❌ Output file not found: {output_file}")

if __name__ == "__main__":
    main() 