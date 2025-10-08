import pandas as pd
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkinter import scrolledtext
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy
import math
import sys
import re
import warnings
warnings.filterwarnings('ignore')
from difflib import SequenceMatcher

class StrongboxParser:
    def __init__(self):
        self.source_file = None
        self.output_dir = None
        self.start_date = None
        self.end_date = None
        self.begin_balance_date = None
        self.source_data = {}
        self.template_data = {}
        self.root = None
        self.status_label = None
        self.progress_bar = None
        self.output_filename = None
        self.console_output = None  # Will hold the text widget for console output
        self.date_columns = {}
        self.presentation_currency = None  # Store presentation currency from TOC tab
        self.non_usd_transactions = []  # Store non-USD transactions
        self.non_usd_headers = ['Journal ID', 'Type', 'Journal Entry Description', 'Posted Date', 'Account ID', 'Journal Line Description', 'Name', 'Debit Amount', 'Credit Amount', 'Transaction Currency']  # Headers for non-USD transactions tab
        self.mapping_categories = {}  # Store mapping categories from Excel file
        self.account_id_map = {}  # Map Original Account Id -> (possibly disambiguated) display Account ID

    def print_and_log(self, message):
        """Print to console and also log to GUI if available"""
        print(message)  # Still print to console
        if self.console_output:
            try:
                self.console_output.insert(tk.END, message + "\n")
                self.console_output.see(tk.END)  # Auto-scroll to bottom
                if self.root:
                    self.root.update()  # Update GUI
            except:
                pass  # Ignore errors if GUI is not available

    def update_status(self, message, progress=None):
        """Update the status message and progress bar"""
        if self.status_label:
            self.status_label.config(text=message)
        if progress is not None and self.progress_bar:
            self.progress_bar['value'] = progress
        if self.root:
            self.root.update()

    def get_file_paths(self):
        """DEPRECATED: Use select_source_file and select_output_location instead"""
        pass

    def determine_date_range(self):
        """Determine date range from TB sheet efficiently"""
        self.update_status('Determining date range from TB sheet...', 10)

        # Read the date row with optimized settings
        try:
            date_row = pd.read_excel(
                self.source_file,
                sheet_name='TB',
                header=None,
                nrows=1,
                skiprows=3
                # Removed pyarrow dependency for better compatibility
            )
            date_row = date_row.iloc[0]

            # Convert all values to datetime efficiently using vectorized operations
            dates = pd.to_datetime(date_row, errors='coerce')
            valid_dates = dates[dates.notna()]

            if valid_dates.empty:
                raise Exception('No valid dates found in TB sheet row 4')
            if len(valid_dates) < 2:
                raise Exception('Need at least two dates in TB sheet')

            # Create date_columns mapping
            date_columns = {date: idx for idx, date in enumerate(dates) if pd.notna(date)}

            # Set date range
            self.begin_balance_date = valid_dates.min()
            self.start_date = self.begin_balance_date + relativedelta(days=1)
            self.end_date = valid_dates.max()
            self.date_columns = date_columns

            self.print_and_log(f'📅 Date range: {self.start_date.strftime("%Y-%m-%d")} to {self.end_date.strftime("%Y-%m-%d")}')
            self.print_and_log(f'📅 Found {len(date_columns)} valid dates in TB sheet')



            return date_columns

        except Exception as e:
            self.print_and_log(f'❌ Error determining date range: {str(e)}')
            raise

    def get_last_day_of_month(self, date):
        """Return the last day of the month for a given date"""
        # If the date is already the last day, just return it
        next_month = date.replace(day=28) + relativedelta(days=4)  # Move to next month
        last_day = next_month - relativedelta(days=next_month.day)  # Subtract days to get last day
        return last_day

    def _initialize_excel_loading(self):
        """Initialize Excel file loading and get sheet names"""
        self.print_and_log("\nStarting to load source data...")
        self.update_status("Loading transaction data...", 40)
        
        self.print_and_log("Opening Excel file...")
        excel_file_pd = pd.ExcelFile(self.source_file)
        self.print_and_log(f"Found sheets: {excel_file_pd.sheet_names}")
        return excel_file_pd

    def _validate_required_tabs(self, excel_file_pd):
        """Validate that all required tabs are present in the Strongbox file"""
        self.print_and_log("\nValidating required tabs...")
        self.update_status("Validating required tabs...", 35)
        
        available_sheets = excel_file_pd.sheet_names
        missing_tabs = []
        
        # Check for TB sheet
        if 'TB' not in available_sheets:
            missing_tabs.append('TB')
        
        # Check for TB-DATA sheet
        if 'TB-DATA' not in available_sheets:
            missing_tabs.append('TB-DATA')
        
        # Check for at least one TXN-FY sheet
        txn_sheets = [sheet for sheet in available_sheets if sheet.startswith('TXN-FY')]
        if not txn_sheets:
            missing_tabs.append('TXN-FY* (at least one transaction sheet starting with TXN-FY)')
        
        # If any required tabs are missing, raise an error
        if missing_tabs:
            if len(missing_tabs) == 1:
                error_message = f"Unable to parse this Strongbox file. Missing required tab: {missing_tabs[0]}"
            else:
                tabs_list = ', '.join(missing_tabs[:-1]) + f' and {missing_tabs[-1]}'
                error_message = f"Unable to parse this Strongbox file. Missing required tabs: {tabs_list}"
            
            self.print_and_log(f"ERROR: {error_message}")
            self.print_and_log(f"Available sheets in file: {available_sheets}")
            self.update_status("Validation failed - missing required tabs", 0)
            raise Exception(error_message)
        
        # Log success
        self.print_and_log(f"✅ All required tabs found:")
        self.print_and_log(f"  • TB sheet: Found")
        self.print_and_log(f"  • TB-DATA sheet: Found") 
        self.print_and_log(f"  • Transaction sheets: Found {len(txn_sheets)} sheets ({', '.join(txn_sheets)})")
        
        return True

    def _try_openpyxl_fallback(self, source_file_path, sheet_to_read, use_data_only):
        """Fallback method to read Excel sheets using openpyxl when pandas fails"""
        self.print_and_log(f"FALLBACK ATTEMPT with data_only={use_data_only} for sheet '{sheet_to_read}'.")
        fallback_workbook = None
        df_from_this_attempt = None
        try:
            fallback_workbook = openpyxl.load_workbook(source_file_path, data_only=use_data_only, read_only=True)
            if sheet_to_read not in fallback_workbook.sheetnames:
                self.print_and_log(f"FALLBACK WARNING: Sheet '{sheet_to_read}' not found in workbook (data_only={use_data_only}).")
                return None

            sheet_obj = fallback_workbook[sheet_to_read]
            data_rows = []
            header = []
            if sheet_obj.max_row > 0:
                header = [cell.value for cell in sheet_obj[1]]
            else:
                self.print_and_log(f"FALLBACK WARNING: Sheet '{sheet_to_read}' (data_only={use_data_only}) appears empty.")

            problematic_cell_count = 0
            for row_idx, row in enumerate(sheet_obj.iter_rows(min_row=2)):
                values = {}
                for col_idx, cell in enumerate(row):
                    cell_value = None
                    try:
                        cell_value = cell.value
                    except Exception as cell_err:
                        self.print_and_log(f"FALLBACK CELL ERROR (data_only={use_data_only}): Sheet '{sheet_to_read}', Row {row_idx + 2}, Col {col_idx + 1}. Error: {cell_err}. Using None.")
                        problematic_cell_count += 1
                    if col_idx < len(header):
                        if header[col_idx] is not None: 
                            values[header[col_idx]] = cell_value
                        else: 
                            values[f"Unknown_Header_Col_{col_idx+1}"] = cell_value 
                    else: 
                        values[f"Extra_Data_Col_{col_idx+1}"] = cell_value
                data_rows.append(values)
            
            df_from_this_attempt = pd.DataFrame(data_rows)
            if df_from_this_attempt.empty and not data_rows and header: 
                df_from_this_attempt = pd.DataFrame(columns=header)
            
            if problematic_cell_count > 0:
                self.print_and_log(f"FALLBACK INFO (data_only={use_data_only}): Encountered {problematic_cell_count} problematic cell(s) in '{sheet_to_read}'.")
            self.print_and_log(f"FALLBACK SUCCESS (data_only={use_data_only}): Successfully read and extracted data for '{sheet_to_read}'.")
            return df_from_this_attempt

        except Exception as current_attempt_err:
            self.print_and_log(f"FALLBACK ERROR (data_only={use_data_only}) for '{sheet_to_read}': {current_attempt_err} (Type: {type(current_attempt_err)})")
            raise current_attempt_err
        finally:
            if fallback_workbook:
                fallback_workbook.close()
                self.print_and_log(f"FALLBACK CLOSED (data_only={use_data_only}): Workbook for '{sheet_to_read}'.")

    def load_source_data(self):
        """Load all required data from the Excel file efficiently using openpyxl directly"""
        self.update_status('Loading transaction data...', 20)

        try:
            # Load workbook once with read-only mode and data_only for better performance
            workbook = openpyxl.load_workbook(self.source_file, read_only=True, data_only=True)

            # Validate required tabs
            available_sheets = workbook.sheetnames
            txn_sheets = [s for s in available_sheets if s.startswith('TXN-FY')]

            # Check for missing required tabs
            missing_tabs = []
            if 'TB' not in available_sheets:
                missing_tabs.append('TB')
            if 'TB-DATA' not in available_sheets:
                missing_tabs.append('TB-DATA')
            if not txn_sheets:
                missing_tabs.append('TXN-FY* (at least one transaction sheet starting with TXN-FY)')

            if missing_tabs:
                if len(missing_tabs) == 1:
                    error_message = f"Unable to parse this Strongbox file. Missing required tab: {missing_tabs[0]}"
                else:
                    tabs_list = ', '.join(missing_tabs[:-1]) + f' and {missing_tabs[-1]}'
                    error_message = f"Unable to parse this Strongbox file. Missing required tabs: {tabs_list}"
                self.print_and_log(f"❌ ERROR: {error_message}")
                self.print_and_log(f"📋 Available sheets in file: {available_sheets}")
                self.update_status("Validation failed - missing required tabs", 0)
                raise Exception(error_message)

            self.print_and_log("✅ All required tabs found:")
            self.print_and_log("  • TB sheet: Found")
            self.print_and_log("  • TB-DATA sheet: Found")
            self.print_and_log(f"  • Transaction sheets: Found {len(txn_sheets)} sheets ({', '.join(txn_sheets)})")

            # Process TXN sheets using pandas for better performance with large datasets
            for sheet_name in txn_sheets:
                try:
                    # Read the sheet with pandas, forcing account-related columns to be strings
                    # to preserve exact format (prevent 1016 -> 1016.0 conversion)
                    dtype_dict = {
                        'Account Id': str,
                        'Account Number/Code': str, 
                        'Account Number': str,
                        'Account Code': str
                    }
                    df = pd.read_excel(self.source_file, sheet_name=sheet_name, dtype=dtype_dict)
                    
                    # Fix account columns to preserve exact decimal formatting (3041.3 should stay 3041.3, not 3041.30)
                    # Remove trailing zeros from decimal account numbers
                    def fix_decimal_formatting(value):
                        if value is None:
                            return ''
                        str_value = str(value).strip()
                        if str_value == '' or str_value.lower() == 'nan':
                            return ''
                        
                        # If it contains a decimal point, remove trailing zeros
                        if '.' in str_value:
                            # Remove trailing zeros after decimal point
                            str_value = str_value.rstrip('0').rstrip('.')
                        
                        return str_value
                    
                    if not df.empty:
                        # Apply decimal formatting fix to account columns
                        account_columns = ['Account Id', 'Account Number/Code', 'Account Number', 'Account Code']
                        for col in account_columns:
                            if col in df.columns:
                                df[col] = df[col].apply(fix_decimal_formatting)
                    
                    # Filter by date range if needed
                    if self.start_date is not None and self.end_date is not None and 'Fiscal Month' in df.columns:
                        # Only include transactions from start_date onwards (not from begin_balance_date)
                        df = df[(df['Fiscal Month'] >= self.start_date) & (df['Fiscal Month'] <= self.end_date)]
                    
                    if not df.empty:
                        # Clean and convert data types
                        df['Transaction Id'] = df['Transaction Id'].astype(str)
                        # Account Id is already a string from dtype specification
                        df['Memo'] = df['Memo'].fillna('')
                        df['Doc/Ref No'] = df['Doc/Ref No'].fillna('')
                        df['Transaction Type'] = df['Transaction Type'].fillna('')
                        df['Relationship Name'] = df['Relationship Name'].fillna('')
                        df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
                        df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
                        
                        # Check for non-USD transactions if presentation currency is not USD
                        self._check_non_usd_transactions(df, sheet_name)
                        
                        # Store the processed data
                        self.source_data[sheet_name] = df
                        self.print_and_log(f"✅ {sheet_name}: {len(df)} transactions loaded")
                    else:
                        self.print_and_log(f"⚠️ {sheet_name}: No transactions in date range")

                except Exception as e:
                    self.print_and_log(f"⚠️ Error loading {sheet_name}: {str(e)}")

            # Load TB-DATA sheet
            self.print_and_log("\nLoading TB-DATA sheet...")
            self.update_status("Loading TB-DATA sheet...", 25)
            try:
                tb_data = pd.read_excel(self.source_file, sheet_name='TB-DATA')
                self.source_data['TB-DATA'] = tb_data
                self.print_and_log(f"✅ Successfully loaded TB-DATA sheet with {len(tb_data)} rows")

            except Exception as e:
                self.print_and_log(f"⚠️ Warning: Could not load TB-DATA sheet: {str(e)}")
                self.print_and_log("Will use default balance values (0) if TB-DATA is not available")
                self.source_data['TB-DATA'] = None

            workbook.close()
            
            # Log summary of non-USD transactions
            if self.non_usd_transactions:
                self.print_and_log(f"\n⚠️ Non-USD transactions found. These are printed on the 'Non-USD Transactions' tab of the output file")
            else:
                self.print_and_log("\n✅ All transactions are in USD")

            # Load trial balance data using openpyxl
            tb_data = self.load_trial_balance_data()
            self.print_and_log(f"Storing TB data in source_data. Shape: {tb_data.shape if tb_data is not None else 'None'}")
            self.print_and_log(f"TB data columns: {tb_data.columns.tolist() if tb_data is not None and not tb_data.empty else 'No columns'}")
            self.source_data['TB'] = tb_data

        except Exception as e:
            self.print_and_log(f"❌ Error loading Excel file: {str(e)}")
            raise
    
    def load_trial_balance_data(self):
        # This method keeps the existing TB sheet loading logic
        self.update_status('Loading trial balance data...', 40)

        # Get the date_columns that were determined in determine_date_range
        date_columns = self.date_columns

        # Find the closest TB dates to our calculated range
        available_tb_dates = sorted(date_columns.keys())

        # Find closest beginning date
        closest_begin_date = None
        for tb_date in available_tb_dates:
            if tb_date <= self.begin_balance_date:
                closest_begin_date = tb_date
            else:
                break

        if closest_begin_date is None:
            closest_begin_date = available_tb_dates[0]

        # Find closest ending date
        closest_end_date = None
        for tb_date in reversed(available_tb_dates):
            if tb_date >= self.end_date:
                closest_end_date = tb_date
            else:
                break

        if closest_end_date is None:
            closest_end_date = available_tb_dates[-1]

        tb_result = self._extract_trial_balance_data(date_columns, closest_begin_date, closest_end_date)
        self.print_and_log(f"load_trial_balance_data() completed. Result shape: {tb_result.shape if tb_result is not None else 'None'}")
        return tb_result

    def create_journal_entries(self):
        """Create Journal Entries & Lines tabs"""
        self.update_status("Creating journal entries...", 60)
        
        # Step 1: Get transaction sheets
        self.print_and_log("\nStep 1: Getting transaction sheets")
        transaction_sheets = {k: v for k, v in self.source_data.items() if k.startswith('TXN-FY')}
        self.print_and_log(f"Found sheets: {list(transaction_sheets.keys())}")
        
        # Step 2: Process each sheet individually
        self.print_and_log("\nStep 2: Processing individual sheets")
        processed_sheets = {}  # Changed from list to dictionary
        for sheet_name, df in transaction_sheets.items():
            try:
                self.print_and_log(f"\nProcessing sheet: {sheet_name}")
                # Create a copy of the dataframe
                df_copy = df.copy()
                
                # Convert columns to appropriate types
                self.print_and_log("Converting columns...")
                df_copy['Transaction Id'] = df_copy['Transaction Id'].astype(str)
                df_copy['Memo'] = df_copy['Memo'].fillna('')
                df_copy['Doc/Ref No'] = df_copy['Doc/Ref No'].fillna('')
                # Account Id is already a string from dtype specification
                
                # Handle optional columns
                for col in ['Transaction Type', 'Relationship Name']:
                    if col in df_copy.columns:
                        df_copy[col] = df_copy[col].fillna('')
                    else:
                        df_copy[col] = ''
                
                # Convert numeric columns
                self.print_and_log("Converting numeric columns...")
                df_copy['Debit'] = pd.to_numeric(df_copy['Debit'], errors='coerce').fillna(0)
                df_copy['Credit'] = pd.to_numeric(df_copy['Credit'], errors='coerce').fillna(0)
                
                # Create the required columns
                self.print_and_log("Creating required columns...")
                
                # Determine Account ID to use: Account Number/Code if available, otherwise Account Name
                # Preserve exact format from input file - don't modify the values
                account_ids = []
                for _, row in df_copy.iterrows():
                    account_number_code = None
                    account_name = None
                    
                    # Check if Account Number/Code column exists and get its value
                    if 'Account Number/Code' in df_copy.columns:
                        account_number_code = row.get('Account Number/Code')
                    elif 'Account Number' in df_copy.columns:
                        account_number_code = row.get('Account Number')
                    elif 'Account Code' in df_copy.columns:
                        account_number_code = row.get('Account Code')
                    
                    # Get Account Name as fallback
                    if 'Account Name' in df_copy.columns:
                        account_name = row.get('Account Name')
                    
                    # Use Account Number/Code if available and not blank, otherwise use Account Name
                    # Preserve original format - just convert to string and check for validity
                    if account_number_code is not None and str(account_number_code).strip() and str(account_number_code).strip().lower() != 'nan':
                        account_ids.append(str(account_number_code).strip())
                    elif account_name is not None and str(account_name).strip() and str(account_name).strip().lower() != 'nan':
                        account_ids.append(str(account_name).strip())
                    else:
                        # Fallback to original Account Id if neither is available, but avoid 'nan'
                        original_account_id = str(row.get('Account Id', '')).strip()
                        if original_account_id and original_account_id.lower() != 'nan':
                            account_ids.append(original_account_id)
                        else:
                            account_ids.append('')  # Use empty string instead of 'nan'
                
                processed_df = pd.DataFrame({
                    'Journal ID': df_copy['Transaction Id'],
                    'Type': df_copy['Transaction Type'],
                    'Journal Entry Description': df_copy['Doc/Ref No'],
                    'Posted Date': df_copy['Transaction Date'],
                    'Account ID': account_ids,
                    'Journal Line Description': df_copy['Memo'],
                    'Name': df_copy['Relationship Name'],
                    'Debit Amount': df_copy['Debit'],
                    'Credit Amount': df_copy['Credit']
                })

                # If available, remap Account ID to the disambiguated ID using Original Account Id from TXN
                try:
                    if 'Account Id' in df_copy.columns:
                        # TXN Account Id corresponds to TB 'Original Account Id'
                        df_copy['Account Id'] = df_copy['Account Id'].astype(str).str.strip()
                        if hasattr(self, 'account_id_map') and self.account_id_map:
                            # Build mapped Account ID per row using Original Account Id
                            mapped_ids = df_copy['Account Id'].map(self.account_id_map)
                            # Where mapping exists, override the Account ID in processed_df
                            processed_df['Account ID'] = mapped_ids.fillna(processed_df['Account ID'])
                except Exception as je_map_err:
                    self.print_and_log(f"WARNING: Failed to map JE Account IDs via Original Account Id: {je_map_err}")
                
                processed_sheets[sheet_name] = processed_df  # Store in dictionary with sheet name as key
                self.print_and_log(f"Successfully processed sheet: {sheet_name}")
                
            except Exception as e:
                self.print_and_log(f"Error processing sheet {sheet_name}: {str(e)}")
                self.print_and_log("DataFrame info:")
                self.print_and_log(df.info())
                self.print_and_log("\nSample data:")
                self.print_and_log(df.head())
                raise
        
        # Step 3: Return dictionary of processed sheets
        self.print_and_log("\nStep 3: Finished processing sheets")
        try:
            if not processed_sheets:
                raise Exception("No sheets were successfully processed")
            
            self.print_and_log(f"Successfully processed {len(processed_sheets)} sheets")
            return processed_sheets
            
        except Exception as e:
            self.print_and_log(f"Error processing sheets: {str(e)}")
            raise

    def create_trial_balance(self):
        """Create Comparative Trial Balances tab"""
        self.update_status("Creating trial balance...", 70)
        tb_data = self.source_data['TB']
        tb_data_data = self.source_data.get('TB-DATA')

        # Only filter out exact header matches, not all rows containing "account"
        if not tb_data.empty and 'Account Id' in tb_data.columns:
            # Convert to string first to handle any non-string values
            tb_data['Account Id'] = tb_data['Account Id'].astype(str)
            tb_data = tb_data[~((tb_data['Account Id'].str.lower() == "account id") | 
                               (tb_data['Account Id'].str.lower() == "account"))]
        else:
            self.print_and_log("⚠️ Warning: TB data is empty or missing 'Account Id' column")
            self.print_and_log(f"TB data columns: {tb_data.columns.tolist() if not tb_data.empty else 'DataFrame is empty'}")
            self.print_and_log(f"TB data shape: {tb_data.shape}")
            if not tb_data.empty:
                self.print_and_log("First few rows of TB data:")
                self.print_and_log(tb_data.head())

        # Get the earliest and latest TB dates from the TB sheet's date row
        tb_dates = sorted(self.date_columns.keys())
        begin_tb_date = tb_dates[0]
        end_tb_date = tb_dates[-1]

        # Build a mapping from (Account Id, Fiscal Month) to Ending Account Balance
        tbdata_lookup = {}
        if tb_data_data is not None:
            for _, row in tb_data_data.iterrows():
                acc_id = str(row['Account Id']) if 'Account Id' in row else str(row[1])
                fiscal_month = pd.to_datetime(row['Fiscal Month']) if pd.notna(row['Fiscal Month']) else None
                if fiscal_month is not None:
                    key = (acc_id, fiscal_month)
                    bal = row['Ending Account Balance'] if 'Ending Account Balance' in row else row.iloc[6]
                    try:
                        bal = float(bal) if pd.notna(bal) else 0.0
                    except (ValueError, TypeError):
                        bal = 0.0
                    tbdata_lookup[key] = bal

        begin_balances = []
        end_balances = []
        if not tb_data.empty and 'Account Id' in tb_data.columns and 'Original Account Id' in tb_data.columns:
            for i, (display_id, original_id) in enumerate(zip(tb_data['Account Id'], tb_data['Original Account Id'])):
                original_id_str = str(original_id)
                begin_bal = tbdata_lookup.get((original_id_str, begin_tb_date), 0.0)
                end_bal = tbdata_lookup.get((original_id_str, end_tb_date), 0.0)
                begin_balances.append(begin_bal)
                end_balances.append(end_bal)
        else:
            self.print_and_log("⚠️ Warning: Cannot process balances - TB data is empty or missing required columns")

        if not tb_data.empty and 'Account Id' in tb_data.columns:
            trial_balance = pd.DataFrame({
                'Account ID': tb_data['Account Id'],
                'Account Name': tb_data['Account Name'] if 'Account Name' in tb_data.columns else [''] * len(tb_data),
                'Beginning Balance \n(Prior Period Balance)': begin_balances,
                'Ending Balance': end_balances,
                'Account Type \n(see Mapping Categories tab)': [''] * len(tb_data),
                'Account Mapping \n(see Mapping Categories tab)': [''] * len(tb_data),
                'Account Description': tb_data['Financial Statement Classification'] if 'Financial Statement Classification' in tb_data.columns else [''] * len(tb_data)
            })
        else:
            # Create empty trial balance if no data
            self.print_and_log("⚠️ Creating empty trial balance due to missing data")
            trial_balance = pd.DataFrame({
                'Account ID': [],
                'Account Name': [],
                'Beginning Balance \n(Prior Period Balance)': [],
                'Ending Balance': [],
                'Account Type \n(see Mapping Categories tab)': [],
                'Account Mapping \n(see Mapping Categories tab)': [],
                'Account Description': []
            })

        # Load mapping categories for automapping
        self.load_mapping_categories()
        
        # Build Account ID disambiguation map for duplicate name-based IDs and apply to TB & later JE
        try:
            tb_df_for_mapping = tb_data[['Account Id', 'Original Account Id', 'Account Name']].copy()
            tb_df_for_mapping['Account Id'] = tb_df_for_mapping['Account Id'].astype(str).str.strip()
            tb_df_for_mapping['Account Name'] = tb_df_for_mapping['Account Name'].astype(str).str.strip()
            tb_df_for_mapping['Original Account Id'] = tb_df_for_mapping['Original Account Id'].astype(str).str.strip()

            # Rows where display Account Id came from Account Name (i.e., number/code blank)
            used_name_mask = (tb_df_for_mapping['Account Name'] != '') & (tb_df_for_mapping['Account Id'] == tb_df_for_mapping['Account Name'])
            duplicate_name_groups = tb_df_for_mapping[used_name_mask].groupby('Account Id').size()
            duplicate_names = set(duplicate_name_groups[duplicate_name_groups > 1].index)

            account_id_map = {}
            # Assign enumerated names for duplicates based on stable order of Original Account Id
            for name in duplicate_names:
                group = tb_df_for_mapping[(used_name_mask) & (tb_df_for_mapping['Account Id'] == name)].copy()
                group = group.sort_values('Original Account Id')
                for i, (_, r) in enumerate(group.iterrows(), start=1):
                    account_id_map[r['Original Account Id']] = f"{name}({i})"

            # Default mapping for all other accounts (including unique names and code-based IDs)
            for _, r in tb_df_for_mapping.iterrows():
                orig = r['Original Account Id']
                if orig not in account_id_map:
                    account_id_map[orig] = r['Account Id']

            # Store map for use in journal entries
            self.account_id_map = account_id_map

            # Apply mapping to the Trial Balance 'Account ID' column
            mapped_ids = tb_df_for_mapping['Original Account Id'].map(self.account_id_map)
            trial_balance['Account ID'] = list(mapped_ids)
        except Exception as map_err:
            self.print_and_log(f"WARNING: Failed to build Account ID disambiguation map: {map_err}")
            try:
                self.account_id_map = {str(oid).strip(): str(aid).strip() for oid, aid in zip(tb_data['Original Account Id'], tb_data['Account Id'])}
            except Exception:
                self.account_id_map = {}

        # Apply the class method to populate the Account Type column
        trial_balance['Account Type \n(see Mapping Categories tab)'] = trial_balance.apply(lambda r: self.determine_account_type(r['Account Description'], r['Account Name']), axis=1)

        # Apply automapping to populate the Account Mapping column
        self.print_and_log("\n🤖 Running automapper to populate Account Mapping column...")
        account_mappings = []
        mapped_count = 0
        
        for _, row in trial_balance.iterrows():
            fs_classification = row['Account Description']
            account_type = row['Account Type \n(see Mapping Categories tab)']
            account_name = row['Account Name']
            
            # Special-case: Non-Operating Income and Expenses -> force nonOperating mapping by account type
            try:
                if isinstance(fs_classification, str) and fs_classification.startswith('Net Income → Total Non-Operating Income and Expenses'):
                    if account_type == 'Income':
                        best_mapping = 'income:nonOperating'
                    elif account_type == 'Expense':
                        best_mapping = 'expense:nonOperating'
                    else:
                        best_mapping = ''
                    if best_mapping:
                        mapped_count += 1
                    account_mappings.append(best_mapping)
                    continue
            except Exception:
                # Fall through to normal mapping if any unexpected issue occurs
                pass

            if account_type and fs_classification:
                best_mapping = self.find_best_mapping(fs_classification, account_type, account_name)
                if best_mapping:
                    mapped_count += 1
                account_mappings.append(best_mapping)
            else:
                account_mappings.append('')
        
        trial_balance['Account Mapping \n(see Mapping Categories tab)'] = account_mappings
        
        self.print_and_log(f"✅ Automapper completed: {mapped_count} of {len(trial_balance)} accounts mapped")

        # Count how many accounts were classified for each type
        account_type_counts = trial_balance['Account Type \n(see Mapping Categories tab)'].value_counts()
        self.print_and_log("\nAccount Type classification summary:")
        for account_type, count in account_type_counts.items():
            if account_type != '':
                self.print_and_log(f"  {account_type}: {count} accounts")

        account_type_col = 'Account Type \n(see Mapping Categories tab)'
        unclassified_count = (trial_balance[account_type_col] == '').sum()
        self.print_and_log(f"  Unclassified: {unclassified_count} accounts")
        
        # Count mapping results
        mapping_col = 'Account Mapping \n(see Mapping Categories tab)'
        unmapped_count = (trial_balance[mapping_col] == '').sum()
        self.print_and_log(f"\nAccount Mapping summary:")
        self.print_and_log(f"  Mapped: {mapped_count} accounts")
        self.print_and_log(f"  Unmapped: {unmapped_count} accounts")



        # Only remove rows that are definitely headers (contain exactly "Account ID")
        headers_to_remove = []
        for idx, row in trial_balance.iterrows():
            account_id = str(row['Account ID']).lower() if pd.notna(row['Account ID']) else ""
            if account_id == "account id" or account_id == "account":
                headers_to_remove.append(idx)
                self.print_and_log(f"Removing header row: {row['Account ID']}")

        if headers_to_remove:
            trial_balance = trial_balance.drop(headers_to_remove)

        # Verify there are no empty account IDs but don't filter out other accounts
        trial_balance = trial_balance[trial_balance['Account ID'].notna() & (trial_balance['Account ID'] != '')]

        # Reset the index after filtering
        trial_balance = trial_balance.reset_index(drop=True)

        return trial_balance

    def _setup_output_file_path(self):
        """Setup the output file path"""
        output_file = os.path.join(
            self.output_dir,
            f"{self.output_filename}_{self.start_date.strftime('%Y%m%d')}_{self.end_date.strftime('%Y%m%d')}.xlsx"
        )
        return output_file

    def _prepare_output_data(self):
        """Prepare trial balance and journal entries data for output"""
        # Load the trial balance data
        self.update_status("Creating trial balance...", 85)
        trial_balance = self.create_trial_balance()
        
        # Process the Journal Entries & Lines
        self.update_status("Creating journal entries...", 90)
        journal_entries_dict = self.create_journal_entries()
        
        # Add accounts from journal entries that are missing from trial balance
        self.print_and_log("\nChecking for accounts in journal entries that are missing from trial balance...")
        self.update_status("Adding missing accounts to trial balance...", 87)
        
        # Get all account IDs in trial balance
        tb_account_ids = set(trial_balance['Account ID'].astype(str))
        
        # Get all unique accounts from all journal entries with their account names
        je_accounts = {}
        for sheet_name, journal_entries in journal_entries_dict.items():
            for _, row in journal_entries.iterrows():
                account_id = str(row['Account ID'])
                if account_id not in je_accounts:
                    je_accounts[account_id] = {
                        'Account ID': account_id,
                        'Account Name': '',  # We don't have account names in journal entries
                        'Beginning Balance \n(Prior Period Balance)': 0.0,
                        'Ending Balance': 0.0,
                        'Account Type \n(see Mapping Categories tab)': '',
                        'Account Mapping \n(see Mapping Categories tab)': '',
                        'Account Description': ''
                    }
        
        # Add missing accounts to trial balance
        missing_accounts = []
        for account_id, account_data in je_accounts.items():
            if account_id not in tb_account_ids and account_id.strip() != '':
                missing_accounts.append(account_data)
        
        if missing_accounts:
            self.print_and_log(f"\nAdding {len(missing_accounts)} missing accounts to trial balance")
            missing_accounts_df = pd.DataFrame(missing_accounts)
            trial_balance = pd.concat([trial_balance, missing_accounts_df], ignore_index=True)
        
        # Clean the data
        trial_balance = self._clean_data_for_excel(trial_balance)
        for sheet_name in journal_entries_dict:
            journal_entries_dict[sheet_name] = self._clean_data_for_excel(journal_entries_dict[sheet_name])
        
        return trial_balance, journal_entries_dict

    def _clean_data_for_excel(self, df):
        """Clean data for Excel output"""
        def ultra_clean_value(value):
            """Clean individual values for Excel"""
            if pd.isna(value):
                return ''
            
            # Convert to string and clean
            str_value = str(value)
            
            # Remove or replace problematic characters
            str_value = str_value.replace('\x00', '')  # Null bytes
            str_value = str_value.replace('\r', ' ')   # Carriage returns
            str_value = str_value.replace('\n', ' ')   # Line feeds
            str_value = str_value.replace('\t', ' ')   # Tabs
            
            # Handle other potential issues
            if len(str_value) > 32767:  # Excel cell character limit
                str_value = str_value[:32767]
            
            return str_value
        
        # Create a copy to avoid modifying the original
        df_clean = df.copy()
        
        # Clean all string columns
        for col in df_clean.columns:
            if df_clean[col].dtype == object:  # Only clean string/object columns
                df_clean[col] = df_clean[col].apply(ultra_clean_value)
        
        return df_clean

    def _create_excel_workbook(self):
        """Create Excel workbook with basic styling setup"""
        self.print_and_log("Creating Excel workbook with openpyxl...")
        workbook = openpyxl.Workbook()
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        blue_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        gray_fill = PatternFill(start_color='999999', end_color='999999', fill_type='solid')
        dark_blue_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        styles = {
            'header_font': header_font,
            'blue_fill': blue_fill,
            'gray_fill': gray_fill,
            'dark_blue_fill': dark_blue_fill,
            'center_alignment': center_alignment
        }
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        return workbook, styles

    def _create_other_sheets(self, workbook, styles):
        """Create and format all other sheets in the specified order"""
        # Create Instructions sheet
        instructions_sheet = workbook.create_sheet('Instructions')
        instructions_sheet.append(['Content for Instructions'])
        
        # Create Data Validation Tests sheet
        validation_sheet = workbook.create_sheet('Data Validation Tests')
        validation_sheet.append(['Content for Data Validation Tests'])
        
        # Create Notes sheet
        notes_sheet = workbook.create_sheet('Notes')
        notes_sheet.append(['Content for Notes'])
        
        # Create Banking Accts sheet with specific formatting
        banking_accts_sheet = workbook.create_sheet('Banking Accts')
        
        # Add headers for Banking Accts
        banking_accts_sheet.append(['Required', 'Required', 'Optional', 'Optional'])
        banking_accts_sheet.append(['Account Number', 'Account Name', 'Institution', 'Currency'])
        
        # Style row 1 headers for Banking Accts
        for col in range(1, 3):  # Columns A-B (Required)
            cell = banking_accts_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['blue_fill']
            cell.alignment = styles['center_alignment']
        
        for col in range(3, 5):  # Columns C-D (Optional)
            cell = banking_accts_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['gray_fill']
            cell.alignment = styles['center_alignment']
        
        # Style row 2 headers for Banking Accts
        for col in range(1, 5):  # All columns A-D
            cell = banking_accts_sheet.cell(row=2, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['dark_blue_fill']
            cell.alignment = styles['center_alignment']
        
        # Create Banking Txn sheet with specific formatting
        banking_txn_sheet = workbook.create_sheet('Banking Txn')
        
        # Add headers for Banking Txn
        banking_txn_sheet.append(['Required', 'Required', 'Required', 'Required'])
        banking_txn_sheet.append(['Posted Date', 'Description', 'Amount', 'Account Number'])
        
        # Style row 1 headers for Banking Txn (all Required)
        for col in range(1, 5):  # Columns A-D (all Required)
            cell = banking_txn_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['blue_fill']
            cell.alignment = styles['center_alignment']
        
        # Style row 2 headers for Banking Txn
        for col in range(1, 5):  # All columns A-D
            cell = banking_txn_sheet.cell(row=2, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['dark_blue_fill']
            cell.alignment = styles['center_alignment']
        
        # Create Mapping Categories sheet and populate with actual mapping data
        mapping_sheet = workbook.create_sheet('Mapping Categories')
        self._populate_mapping_categories_sheet(mapping_sheet, styles)

    def _populate_mapping_categories_sheet(self, mapping_sheet, styles):
        """Populate the Mapping Categories sheet with the built-in mapping data"""
        try:
            # Create structured mapping sheet with separate columns for each account type
            headers = [
                'Account Type',
                'Account Mapping ( Assets)',
                'Account Mapping ( Liabilities)', 
                'Account Mapping (Equity)',
                'Account Mapping (Income)',
                'Account Mapping (Expenses)',
                'Unnamed: 6',
                'Banking Currency Code'
            ]
            
            mapping_sheet.append(headers)
            
            # Style the header row
            for col_idx, header in enumerate(headers, 1):
                cell = mapping_sheet.cell(row=1, column=col_idx)
                cell.font = styles['header_font']
                cell.fill = styles['blue_fill']
                cell.alignment = styles['center_alignment']
            
            # Find the maximum number of mappings to determine row count
            max_mappings = max(len(mappings) for mappings in self.mapping_categories.values()) if self.mapping_categories else 0
            
            # Create rows with account type names in first few rows, then mappings
            account_types = ['Assets', 'Liabilities', 'Equity', 'Income', 'Expense']
            for i, account_type in enumerate(account_types):
                row_data = [''] * len(headers)
                row_data[0] = account_type  # Account Type column
                mapping_sheet.append(row_data)
            
            # Add mapping data in columns
            for row_idx in range(len(account_types) + 1, max_mappings + len(account_types) + 1):
                row_data = [''] * len(headers)
                
                # Add mappings for each account type in their respective columns
                mapping_idx = row_idx - len(account_types) - 1
                
                if mapping_idx < len(self.mapping_categories.get('Assets', [])):
                    row_data[1] = self.mapping_categories['Assets'][mapping_idx]  # Assets column
                
                if mapping_idx < len(self.mapping_categories.get('Liabilities', [])):
                    row_data[2] = self.mapping_categories['Liabilities'][mapping_idx]  # Liabilities column
                
                if mapping_idx < len(self.mapping_categories.get('Equity', [])):
                    row_data[3] = self.mapping_categories['Equity'][mapping_idx]  # Equity column
                
                if mapping_idx < len(self.mapping_categories.get('Income', [])):
                    row_data[4] = self.mapping_categories['Income'][mapping_idx]  # Income column
                
                if mapping_idx < len(self.mapping_categories.get('Expense', [])):
                    row_data[5] = self.mapping_categories['Expense'][mapping_idx]  # Expenses column
                
                mapping_sheet.append(row_data)
            
            # Set column widths
            mapping_sheet.column_dimensions['A'].width = 20  # Account Type
            for col in ['B', 'C', 'D', 'E', 'F']:  # Mapping columns
                mapping_sheet.column_dimensions[col].width = 40
            
            self.print_and_log("✅ Mapping Categories sheet populated with built-in mapping data")
                
        except Exception as e:
            self.print_and_log(f"⚠️ Error populating Mapping Categories sheet: {str(e)}")
            # Create a minimal sheet
            mapping_sheet.append(['Account Type', 'Account Mapping'])
            for col in range(1, 3):
                cell = mapping_sheet.cell(row=1, column=col)
                cell.font = styles['header_font']
                cell.fill = styles['blue_fill']
                cell.alignment = styles['center_alignment']

    def _handle_excel_creation_error(self, e, output_file, trial_balance, journal_entries_dict):
        """Handle errors during Excel file creation"""
        self.print_and_log(f"\nERROR creating Excel file: {str(e)}")
        self.print_and_log("Attempting to save data to CSV files instead...")
        
        try:
            # Save trial balance
            csv_base = os.path.splitext(output_file)[0]
            trial_balance.to_csv(f"{csv_base}_trial_balance.csv", index=False)
            self.print_and_log("✓ Saved trial balance to CSV")
            
            # Save each journal entries sheet
            for sheet_name, journal_entries in journal_entries_dict.items():
                safe_name = sheet_name.replace('/', '_').replace('\\', '_')
                journal_entries.to_csv(f"{csv_base}_journal_entries_{safe_name}.csv", index=False)
                self.print_and_log(f"✓ Saved journal entries from {sheet_name} to CSV")
            
            self.print_and_log("\nData has been saved to CSV files in the same directory.")
            self.print_and_log("Please check these files and try to open them in Excel manually.")
            
        except Exception as csv_error:
            self.print_and_log(f"\nERROR saving to CSV: {str(csv_error)}")
            self.print_and_log("Unable to save data in any format.")
            raise

    def create_output_file(self):
        """Create Excel output file with professional formatting"""
        self.update_status('Creating Excel output...', 80)

        start_str = self.start_date.strftime('%Y%m%d')
        end_str = self.end_date.strftime('%Y%m%d')
        self.output_filename = f'Processed_Strongbox_{start_str}_{end_str}.xlsx'
        
        # Create output path
        output_file = os.path.join(self.output_dir, self.output_filename)

        # Create trial balance and journal entries
        trial_balance = self.create_trial_balance()
        journal_entries_dict = self.create_journal_entries()

        # Post-process: ensure all JE Account IDs exist in Trial Balance
        try:
            trial_balance = self._append_missing_je_accounts_to_trial_balance(trial_balance, journal_entries_dict)
        except Exception as append_err:
            self.print_and_log(f"⚠️ Warning appending missing JE Account IDs to Trial Balance: {append_err}")

        # Clean data for Excel
        trial_balance = self._clean_data_for_excel(trial_balance)
        for sheet_name in journal_entries_dict:
            journal_entries_dict[sheet_name] = self._clean_data_for_excel(journal_entries_dict[sheet_name])

        # Create Excel workbook with proper styling
        workbook, styles = self._create_excel_workbook()

        # Create main data sheets first (in desired order)
        tb_sheet = self._create_trial_balance_sheet(workbook, trial_balance, styles)
        je_sheets = self._create_journal_entries_sheets(workbook, journal_entries_dict, styles)
        
        # Create Non-USD Transactions sheet if needed (only when presentation currency is not USD)
        if self.non_usd_transactions and self.presentation_currency != '(usd) united states dollar':
            self._create_non_usd_transactions_sheet(workbook, styles)

        # Create other sheets after main data sheets
        self._create_other_sheets(workbook, styles)

        # Save workbook
        workbook.save(output_file)
        workbook.close()

        self.print_and_log(f'✅ Output file created: {output_file}')
        return output_file

    def _append_missing_je_accounts_to_trial_balance(self, trial_balance, journal_entries_dict):
        """Append Account IDs from JE sheets that are missing in the Trial Balance.

        - Looks up Account Name from COA tab (Account Number/Code -> Account Name)
        - Sets Beginning/Ending balances to 0.0
        - Guesses Account Type from account name (Assets, Liabilities, Equity, Income, Expense)
        """
        # Collect existing IDs from TB
        tb_ids = set(str(x).strip() for x in trial_balance['Account ID'] if pd.notna(x) and str(x).strip())

        # Collect all Account IDs from JE sheets
        je_ids = set()
        for df in journal_entries_dict.values():
            if 'Account ID' in df.columns:
                for val in df['Account ID']:
                    val_str = str(val).strip()
                    if val_str and val_str.lower() != 'nan':
                        je_ids.add(val_str)

        missing_ids = sorted(list(je_ids - tb_ids))
        if not missing_ids:
            return trial_balance

        self.print_and_log(f"📌 Found {len(missing_ids)} Account IDs in JE not present in Trial Balance. Appending...")

        # Build COA lookup: Account Number/Code -> Account Name
        coa_lookup = {}
        try:
            # Prefer pandas for simplicity; fallback to openpyxl if needed
            coa_df = pd.read_excel(self.source_file, sheet_name='COA', dtype=str)
            # Identify possible number/code columns
            code_cols = [c for c in coa_df.columns if str(c).strip().lower() in ['account number/code', 'account number', 'account code']]
            name_col = None
            for c in coa_df.columns:
                if str(c).strip().lower() == 'account name':
                    name_col = c
                    break
            if code_cols and name_col:
                for _, row in coa_df.iterrows():
                    code_val = None
                    for code_col in code_cols:
                        val = row.get(code_col)
                        if pd.notna(val) and str(val).strip():
                            code_val = str(val).strip()
                            break
                    if code_val is not None:
                        coa_lookup[code_val] = str(row.get(name_col, '')).strip()
        except Exception as e:
            self.print_and_log(f"⚠️ Unable to read COA tab for name lookup: {e}")

        # Prepare rows to append
        append_rows = []
        for account_id in missing_ids:
            account_name = coa_lookup.get(account_id, '')
            # If COA has no name for this code, default name to the Account ID itself
            if not account_name:
                account_name = account_id
            guessed_type = self._guess_account_type_from_name(account_name or account_id)
            append_rows.append({
                'Account ID': account_id,
                'Account Name': account_name,
                'Beginning Balance \n(Prior Period Balance)': 0.0,
                'Ending Balance': 0.0,
                'Account Type \n(see Mapping Categories tab)': guessed_type,
                'Account Mapping \n(see Mapping Categories tab)': '',
                'Account Description': ''
            })

        if append_rows:
            trial_balance = pd.concat([trial_balance, pd.DataFrame(append_rows)], ignore_index=True)
            self.print_and_log(f"✅ Appended {len(append_rows)} accounts to Trial Balance")

        return trial_balance

    def _guess_account_type_from_name(self, account_name):
        """Guess account type from account name using simple keyword heuristics."""
        name = (account_name or '').lower()
        # Income / Revenue
        income_keywords = ['income', 'revenue', 'sales']
        if any(k in name for k in income_keywords):
            return 'Income'

        # Expense / Costs
        expense_keywords = ['expense', 'expenses', 'cost', 'cogs', 'cos', 'rent', 'utilities', 'payroll', 'wages', 'salary', 'supplies', 'fee', 'fees']
        if any(k in name for k in expense_keywords):
            return 'Expense'

        # Assets
        asset_keywords = ['cash', 'bank', 'checking', 'savings', 'ar', 'receivable', 'receivables', 'inventory', 'fixed asset', 'equipment', 'property', 'asset']
        if any(k in name for k in asset_keywords):
            return 'Assets'

        # Liabilities
        liability_keywords = ['ap', 'payable', 'payables', 'loan', 'loans', 'debt', 'note payable', 'mortgage', 'credit card']
        if any(k in name for k in liability_keywords):
            return 'Liabilities'

        # Equity
        equity_keywords = ['equity', 'capital', 'retained earnings', 'member distribution', 'dividend']
        if any(k in name for k in equity_keywords):
            return 'Equity'

        # Default guess
        return 'Expense'

    def process_data(self):
        """Process the data and create output file"""
        try:
            # Currency check - same as notebook
            self.print_and_log('')
            self.print_and_log('💰 CHECKING PRESENTATION CURRENCY...')
            self.print_and_log('=' * 40)
            
            currency_found = False
            currency_message = ""
            
            try:
                self.print_and_log('📋 Opening file to check TOC tab...')
                wb = openpyxl.load_workbook(self.source_file, data_only=True, read_only=True)
                self.print_and_log(f'📋 Available sheets: {wb.sheetnames}')
                
                if 'TOC' in wb.sheetnames:
                    self.print_and_log('📋 TOC tab found, checking cell C10...')
                    toc_sheet = wb['TOC']
                    currency_cell = toc_sheet['C10'].value
                    self.print_and_log(f'📋 Raw value in C10: "{currency_cell}"')
                    
                    if currency_cell:
                        currency = str(currency_cell).strip()
                        self.print_and_log(f'📋 Cleaned currency value: "{currency}"')
                        
                        # Store the presentation currency
                        self.presentation_currency = currency.lower()
                        
                        if currency.lower() == '(usd) united states dollar':
                            currency_message = '✅ PRESENTATION CURRENCY: USD ✅'
                        else:
                            currency_message = f'⚠️ WARNING: PRESENTATION CURRENCY IS NOT USD! ⚠️\n    Found: "{currency}"'
                        currency_found = True
                    else:
                        currency_message = '⚠️ Cell C10 in TOC tab is empty'
                else:
                    currency_message = '⚠️ TOC tab not found in file'
                    
                wb.close()
                
            except Exception as e:
                currency_message = f'⚠️ Error checking currency: {str(e)}'
            
            # Always show the currency result
            self.print_and_log('')
            self.print_and_log(currency_message)
            self.print_and_log('=' * 40)
            self.print_and_log('')
            
            # Determine date range automatically
            self.date_columns = self.determine_date_range()
            
            # Load source data using the determined date range
            self.load_source_data()
            
            # Create output file
            output_file = self.create_output_file()
            
            self.update_status(f"File created successfully at:\n{output_file}", 100)
            messagebox.showinfo("Success", f"File created successfully at:\n{output_file}")
        except Exception as e:
            self.update_status(f"Error: {str(e)}", 0)
            messagebox.showerror("Error", str(e))

    def run(self):
        """Run the parser"""
        try:
            # Create main window
            self.root = tk.Tk()
            self.root.title("Strongbox Parser")
            self.root.geometry("900x700")  # Made wider and taller to accommodate console output

            # Create main frame
            main_frame = ttk.Frame(self.root, padding="20")
            main_frame.pack(fill=tk.BOTH, expand=True)

            # Create title label
            title_label = ttk.Label(main_frame, text="Strongbox Parser", font=("Helvetica", 16, "bold"))
            title_label.pack(pady=20)

            # Create source file selection button
            source_button = ttk.Button(main_frame, text="Select Source File", command=self.select_source_file)
            source_button.pack(pady=20)

            # Create status label
            self.status_label = ttk.Label(main_frame, text="Ready to start...", wraplength=500)
            self.status_label.pack(pady=10)

            # Create progress bar
            self.progress_bar = ttk.Progressbar(main_frame, length=400, mode='determinate')
            self.progress_bar.pack(pady=10)

            # Create console output section
            console_frame = ttk.LabelFrame(main_frame, text="Console Output", padding="10")
            console_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))

            # Create scrolled text widget for console output
            self.console_output = scrolledtext.ScrolledText(
                console_frame, 
                height=20, 
                width=80,
                font=("Consolas", 9),
                state=tk.NORMAL,
                wrap=tk.WORD
            )
            self.console_output.pack(fill=tk.BOTH, expand=True)

            # Add initial message to console
            self.print_and_log("Strongbox Parser initialized. Ready to process files.")

            # Start the main event loop
            self.root.mainloop()
        except Exception as e:
            if self.root:
                self.update_status(f"Error: {str(e)}", 0)
                messagebox.showerror("Error", str(e))
            else:
                messagebox.showerror("Error", str(e))
                
    def select_source_file(self):
        """Select source file using file dialog"""
        self.update_status("Selecting source file...", 10)
        self.source_file = filedialog.askopenfilename(
            title="Select Strongbox File",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not self.source_file:
            self.update_status("Operation cancelled: No source file selected", 0)
            return
            
        self.update_status(f"Selected source file: {os.path.basename(self.source_file)}", 20)
        # Proceed to next step - select output file
        self.select_output_location()
    
    def select_output_location(self):
        """Select output directory and filename at once using save file dialog"""
        self.update_status("Selecting output location...", 30)
        
        # Use a save file dialog to get both the directory and filename at once
        output_file = filedialog.asksaveasfilename(
            title="Save Output File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Audit_Sight_Output.xlsx"
        )
        
        if not output_file:
            self.update_status("Operation cancelled: No output location selected", 0)
            return
            
        # Split the output_file into directory and filename
        self.output_dir = os.path.dirname(output_file)
        self.output_filename = os.path.splitext(os.path.basename(output_file))[0]
        
        self.update_status(f"Selected output: {self.output_filename} in {self.output_dir}", 35)
        
        # Proceed to determine date range and process data
        try:
            # Process the data
            self.process_data()
        except Exception as e:
            self.update_status(f"Error: {str(e)}", 0)
            messagebox.showerror("Error", str(e))

    def load_mapping_categories(self):
        """Load built-in mapping categories for automapping"""
        self.print_and_log("📋 Loading built-in mapping categories for automapping...")
        
        # Built-in mapping categories - these are the standard options for each account type
        self.mapping_categories = {
            "Assets": [
                "asset:current:cash",
                "asset:current:undepositedFunds",
                "asset:current:cashNonBankAccounts",
                "asset:current:cashDepositsandInvestments",
                "asset:current:investments",
                "asset:current:accountsReceivable",
                "asset:current:accountsReceivable:allowanceforDoubtfulAccounts",
                "asset:current:inventory",
                "asset:current:prepaidExpenses",
                "asset:current:interCompany",
                "asset:current:other",
                "asset:nonCurrent:fixed",
                "asset:nonCurrent:intangibles",
                "asset:nonCurrent:goodwill",
                "asset:nonCurrent:deposits",
                "asset:nonCurrent:interCompany",
                "asset:nonCurrent:other",
            ],
            "Liabilities": [
                "liability:current:accountsPayable",
                "liability:current:creditCardPayable",
                "liability:current:interCompany",
                "liability:current:payrollLiability",
                "liability:current:salesTax",
                "liability:current:tax",
                "liability:current:other:accrued",
                "liability:current:other",
                "liability:current:other:deferredRevenue",
                "liability:current:other:customerDeposits",
                "liability:current:other:billingsInExcessOfCost",
                "liability:current:debt",
                "liability:noncurrent:debt",
                "liability:noncurrent:tax",
                "liability:noncurrent:interCompany",
                "liability:noncurrent:other",
            ],
            "Equity": [
                "equity:ownersEquity",
                "equity:retainedEarnings",
                "equity:netIncome",
            ],
            "Income": [
                "income:operating",
                "income:nonOperating",
            ],
            "Expense": [
                "expense:costOfGoodsSold",
                "expense:costOfGoodsSold:labor",
                "expense:costOfGoodsSold:materials",
                "expense:costOfGoodsSold:other",
                "expense:operating",
                "expense:operating:advertisingAndMarketing",
                "expense:operating:amortization",
                "expense:operating:bankAndCreditCardFees",
                "expense:operating:depreciation",
                "expense:operating:legalAndProfessionalFees",
                "expense:operating:mealsAndEntertainment",
                "expense:operating:office",
                "expense:operating:payroll",
                "expense:operating:rent",
                "expense:operating:software",
                "expense:operating:travel",
                "expense:operating:other",
                "expense:nonOperating",
                "expense:nonOperating:interest",
                "expense:nonOperating:taxes",
            ],
        }
        
        self.print_and_log(f"✅ Loaded built-in mapping categories:")
        for account_type, mappings in self.mapping_categories.items():
            self.print_and_log(f"  • {account_type}: {len(mappings)} options")

    def find_best_mapping(self, fs_classification, account_type, account_name=''):
        """Find the best mapping option based on Financial Statement Classification Path, Account Type, and Account Name"""
        if not self.mapping_categories or account_type not in self.mapping_categories:
            return ''
        
        if pd.isna(fs_classification) or fs_classification == '':
            return ''
        
        fs_classification = str(fs_classification).strip().lower()
        account_name = str(account_name).strip().lower() if account_name else ''
        available_mappings = self.mapping_categories[account_type]
        
        if not available_mappings:
            return ''
        
        # Extract keywords from both financial statement classification and account name
        fs_keywords = self._extract_keywords(fs_classification)
        name_keywords = self._extract_keywords(account_name) if account_name else []
        
        # Combine keywords, giving more weight to account name keywords
        all_keywords = fs_keywords + name_keywords
        
        best_match = ''
        best_score = 0.0
        
        for mapping in available_mappings:
            mapping_lower = mapping.lower()
            score = self._calculate_mapping_score(fs_keywords, name_keywords, mapping_lower, fs_classification, account_name)
            
            if score > best_score:
                best_score = score
                best_match = mapping
        
        # Apply default mappings if no good match found
        if best_score <= 0.1:
            default_mapping = self._get_default_mapping(fs_classification, account_name, account_type)
            if default_mapping:
                return default_mapping
        
        # Only return a match if the score is above a threshold
        if best_score > 0.1:  # Minimum 10% match
            return best_match
        
        return ''

    def _extract_keywords(self, fs_classification):
        """Extract meaningful keywords from Financial Statement Classification Path"""
        # Remove common path separators and words
        text = fs_classification.replace('→', ' ').replace('total', '').replace('net', '')
        
        # Split into words and filter out common words
        words = re.findall(r'\b\w+\b', text.lower())
        
        # Filter out very common words that don't help with classification
        stop_words = {'and', 'or', 'the', 'of', 'in', 'to', 'for', 'with', 'by', 'from', 'on', 'at', 'as'}
        keywords = [word for word in words if len(word) > 2 and word not in stop_words]
        
        return keywords

    def _calculate_mapping_score(self, fs_keywords, name_keywords, mapping_lower, fs_classification, account_name):
        """Calculate a score for how well a mapping matches the financial statement classification and account name"""
        score = 0.0
        
        # Direct keyword matches from financial statement classification
        for keyword in fs_keywords:
            if keyword in mapping_lower:
                score += 1.0
        
        # Direct keyword matches from account name (higher weight)
        for keyword in name_keywords:
            if keyword in mapping_lower:
                score += 1.5  # Account name gets higher priority
        
        # Partial matches using SequenceMatcher
        fs_similarity = SequenceMatcher(None, fs_classification, mapping_lower).ratio()
        score += fs_similarity * 0.5
        
        if account_name:
            name_similarity = SequenceMatcher(None, account_name, mapping_lower).ratio()
            score += name_similarity * 0.7  # Account name similarity gets higher weight
        
        # Specific business logic mappings
        score += self._apply_business_logic_scoring(fs_keywords, name_keywords, mapping_lower, fs_classification, account_name)
        
        return score

    def _apply_business_logic_scoring(self, fs_keywords, name_keywords, mapping_lower, fs_classification, account_name):
        """Apply business logic for specific account type mappings"""
        score = 0.0
        all_keywords = fs_keywords + name_keywords
        combined_text = f"{fs_classification} {account_name}".lower()
        
        # Cash-related mappings - bank accounts should map to asset:current:cash
        if any(word in all_keywords for word in ['cash', 'bank', 'checking', 'savings']):
            if mapping_lower == 'asset:current:cash':
                score += 5.0  # Higher priority for basic cash mapping
            elif 'cashnonbankaccounts' in mapping_lower.replace(':', ''):
                score -= 2.0  # Strong penalty for non-bank cash accounts when bank terms are present
            elif 'cash' in mapping_lower:
                score += 1.0  # Lower score for other cash mappings
        
        # Specifically for bank accounts - boost basic cash mapping
        if any(word in all_keywords for word in ['bank']) and 'accounts' in combined_text:
            if mapping_lower == 'asset:current:cash':
                score += 3.0  # Extra boost for bank accounts -> basic cash
        
        # Accounts receivable
        if any(word in all_keywords for word in ['receivable', 'receivables', 'ar']):
            if 'accountsreceivable' in mapping_lower.replace(':', ''):
                score += 2.0
        
        # Accounts payable
        if any(word in all_keywords for word in ['payable', 'payables', 'ap']):
            if 'accountspayable' in mapping_lower.replace(':', ''):
                score += 2.0
        
        # Inventory
        if 'inventory' in all_keywords:
            if 'inventory' in mapping_lower:
                score += 2.0
        
        # Fixed assets and depreciation - improved logic
        if any(word in all_keywords for word in ['fixed', 'equipment', 'property', 'plant', 'depreciation', 'depletion', 'amortization', 'impairment']):
            if 'fixed' in mapping_lower:
                score += 2.0
        
        # Other current assets
        if 'other' in combined_text and 'current' in combined_text and 'assets' in combined_text:
            if mapping_lower == 'asset:current:other':
                score += 3.0
        
        # Other current liabilities
        if 'other' in combined_text and 'current' in combined_text and 'liabilities' in combined_text:
            if mapping_lower == 'liability:current:other':
                score += 3.0
        
        # Credit card debt mapping
        if any(word in all_keywords for word in ['credit', 'card']):
            if mapping_lower == 'liability:current:creditcardpayable':
                score += 3.0
        
        # Payroll liabilities mapping
        if any(word in all_keywords for word in ['payroll']) and any(word in all_keywords for word in ['liabilities', 'liability']):
            if mapping_lower == 'liability:current:payrollliability':
                score += 3.0
        
        # Term loans mapping
        if 'term' in all_keywords and any(word in all_keywords for word in ['loan', 'loans']):
            if mapping_lower == 'liability:noncurrent:debt':
                score += 3.0
        
        # Debt mappings - improved specificity
        debt_keywords = ['loan', 'debt', 'note', 'borrowing', 'shrhlder', 'shareholder']
        if any(word in all_keywords for word in debt_keywords):
            if 'current' in combined_text and 'debt' in mapping_lower:
                score += 2.5
            elif 'noncurrent' in mapping_lower and 'debt' in mapping_lower:
                score += 2.5
            elif 'term' in combined_text and 'noncurrent' in mapping_lower and 'debt' in mapping_lower:
                score += 3.0
        else:
            # Penalize debt mappings for non-debt accounts
            if 'debt' in mapping_lower:
                score -= 1.0
        
        # Specific shareholder loan mapping
        if any(word in all_keywords for word in ['shrhlder', 'shareholder']) and 'loan' in combined_text:
            if 'current' in combined_text and mapping_lower == 'liability:current:debt':
                score += 4.0
            elif 'noncurrent' in combined_text and mapping_lower == 'liability:noncurrent:debt':
                score += 4.0
        
        # Revenue/Sales
        if any(word in all_keywords for word in ['sales', 'revenue']):
            if 'income' in mapping_lower and 'operating' in mapping_lower:
                score += 2.0
        
        # Cost of Goods Sold
        if any(word in all_keywords for word in ['cogs', 'cos', 'cost']):
            if 'costofgoodssold' in mapping_lower.replace(':', ''):
                score += 2.0
        
        # Payroll
        if any(word in all_keywords for word in ['payroll', 'wages', 'salary']):
            if 'payroll' in mapping_lower:
                score += 2.0
        
        # Rent
        if 'rent' in all_keywords:
            if 'rent' in mapping_lower:
                score += 2.0
        
        # Interest - more specific matching to avoid false positives
        if any(word in all_keywords for word in ['interest']):
            # Check if it's an operating expense interest (should be non-operating)
            if 'operating' in combined_text and 'expenses' in combined_text:
                if mapping_lower == 'expense:nonoperating:interest':
                    score += 3.0  # Strong boost for operating expenses interest -> non-operating interest
            # Only map to interest if it's clearly loan/debt related
            elif any(word in all_keywords for word in ['loan', 'debt', 'borrowing', 'finance']) or 'interest expense' in combined_text.lower():
                if 'interest' in mapping_lower:
                    score += 2.0
            else:
                # Penalize interest mapping for non-loan related accounts
                if 'interest' in mapping_lower:
                    score -= 2.0
                # If account has "interest" but isn't loan-related, favor generic operating over specific categories
                if 'operating' in mapping_lower and mapping_lower == 'expense:operating':
                    score += 1.0  # Boost generic operating for ambiguous cases
        
        # Bank charges and fees
        if any(word in all_keywords for word in ['bank', 'charge', 'fee']) and any(word in all_keywords for word in ['charge', 'fee']):
            if 'bankandcreditcardfees' in mapping_lower.replace(':', ''):
                score += 2.5
        
        # Auto expenses
        if 'auto' in all_keywords:
            if mapping_lower == 'expense:operating':
                score += 2.0
        
        # Taxes - only boost tax mappings for accounts that clearly mention tax
        if any(word in all_keywords for word in ['tax', 'taxes']):
            if 'tax' in mapping_lower:
                score += 2.0
        else:
            # Penalize tax mappings for accounts that don't mention tax
            if 'tax' in mapping_lower:
                score -= 1.5
        
        # Prefer base operating categories for generic accounts
        if not any(word in all_keywords for word in ['cash', 'bank', 'receivable', 'payable', 'inventory', 'tax', 'payroll', 'debt', 'loan']):
            if mapping_lower in ['expense:operating', 'income:operating']:
                score += 2.0  # Strong boost for base operating categories for generic accounts
            elif mapping_lower.endswith(':other'):
                score += 1.0  # Moderate boost for "other" categories for generic accounts
        
        return score

    def _get_default_mapping(self, fs_classification, account_name, account_type):
        """Get default mapping when no good match is found"""
        combined_text = f"{fs_classification} {account_name}".lower()
        
        # Default expense mapping
        if account_type == 'Expense':
            return 'expense:operating'
        
        # Default income mapping
        if account_type == 'Income':
            return 'income:operating'
        
        # Default asset mapping - use more general categories
        if account_type == 'Assets':
            return 'asset:current:other'  # Default to current:other for assets
        
        # Default liability mapping - use more general categories
        if account_type == 'Liabilities':
            return 'liability:current:other'  # Default to current:other for liabilities
        
        return ''

    def determine_account_type(self, fs_classification):
        """Determine Account Type based on Financial Statement Classification Path"""
        if pd.isna(fs_classification) or fs_classification == '':
            return ''
            
        fs_classification = str(fs_classification).strip()
        
        if fs_classification.startswith('Total Assets'):
            return 'Assets'
        elif fs_classification.startswith('Total Liabilities and Equity → Total Liabilities'):
            return 'Liabilities'
        elif fs_classification.startswith('Total Liabilities and Equity → Total Equity'):
            return 'Equity'
        elif fs_classification.startswith('Net Income → Operating Profit → Gross Profit → Total Net Sales'):
            return 'Income'
        elif fs_classification.startswith('Net Income → Operating Profit → Gross Profit → Total COGS/COS'):
            return 'Expense'
        elif fs_classification.startswith('Net Income → Operating Profit → Total Operating Expenses'):
            return 'Expense'
        else:
            return ''

    def create_test_file(self):
        """Create a minimal test Excel file to diagnose corruption issues"""
        output_file = os.path.join(
            self.output_dir,
            f"TEST_{self.output_filename}_{self.start_date.strftime('%Y%m%d')}_{self.end_date.strftime('%Y%m%d')}.xlsx"
        )
        
        # Create minimal test data
        test_data = [
            ['Account ID', 'Account Name', 'Beginning Balance', 'Ending Balance'],
            ['100', 'Cash', 1000.00, 1500.00],
            ['200', 'Accounts Receivable', 5000.00, 4500.00],
            ['300', 'Inventory', 2000.00, 2200.00]
        ]
        
        # Try the most basic Excel creation possible
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Test Data'
            
            for row_data in test_data:
                sheet.append(row_data)
            
            workbook.save(output_file)
            self.print_and_log(f"Test file created: {output_file}")
            return output_file
        except Exception as e:
            self.print_and_log(f"Error creating test file: {str(e)}")
            return None

    def _apply_data_type_conversions(self, df, sheet_name):
        """Apply data type conversions to DataFrame columns"""
        self.print_and_log(f"Converting data types for {sheet_name}...")
        
        # Convert string columns
        for col in ['Transaction Id', 'Account Id', 'Memo', 'Doc/Ref No']:
            if col in df.columns:
                df[col] = df[col].astype(str)
            else:
                self.print_and_log(f"Warning: Column '{col}' not found in sheet {sheet_name}. Skipping conversion.")
        
        # Convert numeric columns
        for col in ['Debit', 'Credit']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            else:
                self.print_and_log(f"Warning: Column '{col}' not found in sheet {sheet_name}. Skipping conversion.")
        
        return df

    def _check_non_usd_transactions(self, df, sheet_name):
        """Check for non-USD transactions and store them separately only if presentation currency is not USD"""
        if 'Transaction Currency' not in df.columns:
            self.print_and_log(f"INFO: No 'Transaction Currency' column found in {sheet_name}. Assuming all transactions are USD.")
            return
        
        # Only separate non-USD transactions if presentation currency is not USD
        if self.presentation_currency != '(usd) united states dollar':
            self.print_and_log(f"INFO: Presentation currency is not USD, separating non-USD transactions in {sheet_name}")
            
            # Check for non-USD transactions
            non_usd_mask = (df['Transaction Currency'].notna()) & (df['Transaction Currency'].str.upper() != 'USD')
            non_usd_df = df[non_usd_mask]
            
            if len(non_usd_df) > 0:
                self.print_and_log(f"⚠️ Found {len(non_usd_df)} non-USD transactions in {sheet_name}")
                
                # Store non-USD transactions for the separate tab
                for _, row in non_usd_df.iterrows():
                    # Determine Account ID to use: Account Number/Code if available, otherwise Account Name
                    account_number_code = None
                    account_name = None
                    
                    # Check if Account Number/Code column exists and get its value
                    if 'Account Number/Code' in df.columns:
                        account_number_code = row.get('Account Number/Code')
                    elif 'Account Number' in df.columns:
                        account_number_code = row.get('Account Number')
                    elif 'Account Code' in df.columns:
                        account_number_code = row.get('Account Code')
                    
                    # Get Account Name as fallback
                    if 'Account Name' in df.columns:
                        account_name = row.get('Account Name')
                    
                    # Use Account Number/Code if available and not blank, otherwise use Account Name
                    # Preserve original format - just convert to string and check for validity
                    if account_number_code is not None and str(account_number_code).strip() and str(account_number_code).strip().lower() != 'nan':
                        account_id = str(account_number_code).strip()
                    elif account_name is not None and str(account_name).strip() and str(account_name).strip().lower() != 'nan':
                        account_id = str(account_name).strip()
                    else:
                        # Fallback to original Account Id if neither is available, but avoid 'nan'
                        original_account_id = str(row.get('Account Id', '')).strip()
                        if original_account_id and original_account_id.lower() != 'nan':
                            account_id = original_account_id
                        else:
                            account_id = ''  # Use empty string instead of 'nan'
                    
                    non_usd_data = {
                        'Journal ID': str(row.get('Transaction Id', '')),
                        'Type': str(row.get('Transaction Type', '')),
                        'Journal Entry Description': str(row.get('Doc/Ref No', '')),
                        'Posted Date': row.get('Transaction Date', ''),
                        'Account ID': account_id,
                        'Journal Line Description': str(row.get('Memo', '')),
                        'Name': str(row.get('Relationship Name', '')),
                        'Debit Amount': float(row.get('Debit', 0)),
                        'Credit Amount': float(row.get('Credit', 0)),
                        'Transaction Currency': str(row.get('Transaction Currency', ''))
                    }
                    self.non_usd_transactions.append(non_usd_data)
                
                # Remove non-USD transactions from the main dataframe
                df.drop(non_usd_df.index, inplace=True)
                self.print_and_log(f"INFO: Removed {len(non_usd_df)} non-USD transactions from {sheet_name}. Remaining transactions: {len(df)}")
            else:
                self.print_and_log(f"INFO: All transactions in {sheet_name} are USD.")
        else:
            self.print_and_log(f"INFO: Presentation currency is USD, including all transactions in main output for {sheet_name}")

    def _process_date_columns(self, df, sheet_name):
        """Process and validate date columns in the DataFrame"""
        # Check if both Transaction Date and Fiscal Month columns exist
        if 'Transaction Date' not in df.columns:
            self.print_and_log(f"WARNING: 'Transaction Date' column NOT FOUND in sheet {sheet_name}. Cannot filter by date.")
            self.print_and_log(f"Skipping sheet {sheet_name} due to missing 'Transaction Date' column.")
            return None
            
        # Process the Transaction Date and Fiscal Month columns
        self.print_and_log(f"INFO: 'Transaction Date' column found in {sheet_name}.")
        self.print_and_log(f"Sample raw 'Transaction Date' values in {sheet_name} before pd.to_datetime:")
        self.print_and_log(df['Transaction Date'].head(10) if len(df) > 0 else "Sheet is empty or has no dates")
        
        # Convert Transaction Date to datetime
        df['Transaction Date'] = pd.to_datetime(df['Transaction Date'], errors='coerce')
        
        # Check if Fiscal Month column exists for date validation
        has_fiscal_month = 'Fiscal Month' in df.columns
        
        if has_fiscal_month:
            self.print_and_log(f"INFO: 'Fiscal Month' column found in {sheet_name}.")
            self.print_and_log(f"Sample raw 'Fiscal Month' values in {sheet_name} before pd.to_datetime:")
            self.print_and_log(df['Fiscal Month'].head(10) if len(df) > 0 else "Sheet is empty or has no fiscal months")
            
            # Convert Fiscal Month to datetime for comparison
            df['Fiscal Month'] = pd.to_datetime(df['Fiscal Month'], errors='coerce')
            
            # Count NaT values in Fiscal Month
            fiscal_nat_count = df['Fiscal Month'].isnull().sum()
            self.print_and_log(f"INFO: Found {fiscal_nat_count} NaT values in 'Fiscal Month' for {sheet_name} after conversion.")
        
        # Count NaT values in Transaction Date
        transaction_nat_count = df['Transaction Date'].isnull().sum()
        self.print_and_log(f"INFO: Found {transaction_nat_count} NaT values in 'Transaction Date' for {sheet_name} after conversion.")
        
        # Filter out rows where date conversion failed (NaT) for Transaction Date
        df = df[pd.notna(df['Transaction Date'])]
        if has_fiscal_month:
            # Keep rows only where Fiscal Month is valid
            df = df[pd.notna(df['Fiscal Month'])]
        
        self.print_and_log(f"INFO: Rows in {sheet_name} after filtering NaT dates: {len(df)}")
        return df, has_fiscal_month

    def _apply_date_range_filtering(self, df, sheet_name, has_fiscal_month):
        """Apply date range filtering and adjust transaction dates if needed"""
        if len(df) == 0:
            self.print_and_log(f"INFO: No valid dates found in {sheet_name} after NaT filtering. Sheet will not be in final output.")
            return None
        
        # Check if the sheet has any rows in our date range
        # Use Fiscal Month for range filtering if available
        if has_fiscal_month:
            # Filter based on Fiscal Month for date range eligibility
            fiscal_month_mask = (df['Fiscal Month'] >= self.start_date) & (df['Fiscal Month'] <= self.end_date)
            df_in_range = df[fiscal_month_mask].copy()
            self.print_and_log(f"INFO: Rows in {sheet_name} with Fiscal Month in date range [{self.start_date.strftime('%Y-%m-%d')} - {self.end_date.strftime('%Y-%m-%d')}]: {len(df_in_range)}")
            
            if len(df_in_range) > 0:
                df_in_range = self._adjust_transaction_dates(df_in_range, sheet_name)
                return df_in_range
            else:
                self.print_and_log(f"INFO: No data from {sheet_name} within the specified date range based on Fiscal Month. Sheet will not be in final output.")
                return None
        else:
            # If no Fiscal Month, fall back to Transaction Date filtering
            transaction_date_mask = (df['Transaction Date'] >= self.start_date) & (df['Transaction Date'] <= self.end_date)
            df_filtered = df[transaction_date_mask]
            self.print_and_log(f"INFO: Rows in {sheet_name} after applying date range to Transaction Date [{self.start_date.strftime('%Y-%m-%d')} - {self.end_date.strftime('%Y-%m-%d')}]: {len(df_filtered)}")
            
            if len(df_filtered) > 0:
                return df_filtered
            else:
                self.print_and_log(f"INFO: No data from {sheet_name} within the specified date range based on Transaction Date. Sheet will not be in final output.")
                return None

    def _adjust_transaction_dates(self, df_in_range, sheet_name):
        """Adjust transaction dates that don't match their fiscal month"""
        # Identify and fix Transaction Dates that don't match their Fiscal Month
        mismatched_dates = 0
        adjusted_dates = []
        
        for idx, row in df_in_range.iterrows():
            transaction_date = row['Transaction Date']
            fiscal_month = row['Fiscal Month']
            
            # Check if Transaction Date's month/year matches Fiscal Month's month/year
            if (transaction_date.year != fiscal_month.year) or (transaction_date.month != fiscal_month.month):
                # Get the last day of the Fiscal Month
                last_day_of_month = self.get_last_day_of_month(fiscal_month)
                adjusted_dates.append((idx, transaction_date, last_day_of_month))
                # Update Transaction Date to the last day of the Fiscal Month
                df_in_range.at[idx, 'Transaction Date'] = last_day_of_month
                mismatched_dates += 1
        
        if mismatched_dates > 0:
            self.print_and_log(f"INFO: Adjusted {mismatched_dates} Transaction Dates to match their Fiscal Month in {sheet_name}")
            if len(adjusted_dates) > 0 and len(adjusted_dates) <= 10:
                self.print_and_log("Sample of adjusted dates (idx, original_date, new_date):")
                for adj in adjusted_dates[:10]:
                    self.print_and_log(f"  Row {adj[0]}: {adj[1].strftime('%Y-%m-%d')} -> {adj[2].strftime('%Y-%m-%d')}")
        
        return df_in_range

    def _load_trial_balance_data(self):
        """Load trial balance data from TB sheet"""
        self.print_and_log("\nLoading trial balance data...")
        self.update_status("Loading trial balance data...", 50)
        
        # Get the date_columns that were determined in determine_date_range
        date_columns = self.date_columns
        
        # Find the closest TB dates to our calculated range
        available_tb_dates = sorted(date_columns.keys())
        
        # Find closest beginning date (on or before our begin_balance_date)
        closest_begin_date = None
        for tb_date in available_tb_dates:
            if tb_date <= self.begin_balance_date:
                closest_begin_date = tb_date
            else:
                break
        
        # If no TB date is on or before our begin_balance_date, use the earliest TB date
        if closest_begin_date is None:
            closest_begin_date = available_tb_dates[0]
            self.print_and_log(f"WARNING: No TB date found on or before calculated begin balance date {self.begin_balance_date.strftime('%Y-%m-%d')}, using earliest TB date {closest_begin_date.strftime('%Y-%m-%d')}")
        
        # Find closest ending date (on or after our end_date)
        closest_end_date = None
        for tb_date in reversed(available_tb_dates):
            if tb_date >= self.end_date:
                closest_end_date = tb_date
            else:
                break
        
        # If no TB date is on or after our end_date, use the latest TB date
        if closest_end_date is None:
            closest_end_date = available_tb_dates[-1]
            self.print_and_log(f"WARNING: No TB date found on or after calculated end date {self.end_date.strftime('%Y-%m-%d')}, using latest TB date {closest_end_date.strftime('%Y-%m-%d')}")
        
        # Use the beginning and ending dates already identified
        begin_date = self.begin_balance_date
        
        self.print_and_log(f"\nTarget dates:")
        self.print_and_log(f"Begin date: {begin_date}")
        self.print_and_log(f"End date: {self.end_date}")
        self.print_and_log(f"Begin balance date (exact): {closest_begin_date}")
        self.print_and_log(f"End balance date (exact): {closest_end_date}")
        
        return self._extract_trial_balance_data(date_columns, closest_begin_date, closest_end_date)

    def _extract_trial_balance_data(self, date_columns, closest_begin_date, closest_end_date):
        """Extract data from the TB sheet using openpyxl"""
        try:
            self.print_and_log("Opening workbook for trial balance data...")
            self.update_status("Opening workbook for trial balance...", 52)
            wb = openpyxl.load_workbook(self.source_file, data_only=True, read_only=True)
            self.print_and_log("Workbook opened successfully")
            
            self.print_and_log("Accessing TB sheet...")
            self.update_status("Accessing TB sheet...", 54)
            tb_sheet = wb['TB']
            self.print_and_log(f"TB sheet accessed. Sheet has {tb_sheet.max_row} rows and {tb_sheet.max_column} columns")
            
            # Find Financial Statement Classification Path column
            self.print_and_log("Finding Financial Statement Classification column...")
            self.update_status("Finding classification column...", 56)
            fin_statement_col = self._find_financial_classification_column(tb_sheet)
            
            # Print column indices for debugging
            self.print_and_log(f"\nColumn indices:")
            self.print_and_log(f"Begin date column index: {date_columns[closest_begin_date]}")
            self.print_and_log(f"End date column index: {date_columns[closest_end_date]}")
            self.print_and_log(f"Financial Statement Classification column index: {fin_statement_col}")
            
            self.print_and_log("Starting to extract data from TB sheet...")
            self.update_status("Extracting trial balance data...", 58)
            
            data = []
            rows_processed = 0
            
            # Limit the number of rows to process to avoid hanging on very large sheets
            max_row_to_process = min(tb_sheet.max_row, 10000)  # Limit to 10,000 rows max
            
            for row_idx in range(8, max_row_to_process + 1):
                rows_processed += 1
                
                # Update progress every 1000 rows
                if rows_processed % 1000 == 0:
                    progress = 58 + (rows_processed / max_row_to_process) * 10  # 58-68%
                    self.update_status(f"Processing TB rows... ({rows_processed}/{max_row_to_process})", progress)
                
                try:
                    # Get Account Number/Code from column 5, fallback to Account Name from column 6
                    account_number_code = tb_sheet.cell(row=row_idx, column=5).value
                    account_name = tb_sheet.cell(row=row_idx, column=6).value
                    
                    # Get the original Account ID from column 4 (needed for TB-DATA lookup)
                    original_account_id = tb_sheet.cell(row=row_idx, column=4).value
                    
                    # Use Account Number/Code if available, otherwise use Account Name for DISPLAY
                    # Preserve exact format from input file - don't modify the values
                    if account_number_code is not None and str(account_number_code).strip() and str(account_number_code).strip() != 'None':
                        display_account_id = str(account_number_code).strip()
                    elif account_name is not None and str(account_name).strip():
                        display_account_id = str(account_name).strip()
                    else:
                        # Skip rows with no account identifier
                        continue
                    
                    # Skip if original Account ID is also missing (needed for TB-DATA lookup)
                    if original_account_id is None:
                        continue
                    
                    original_account_id = str(original_account_id).strip()
                    
                    # Only skip rows with exact header matches, including the new header format
                    if display_account_id.lower() in ['account id', 'account', 'account number/code']:
                        continue
                        
                    begin_cell = tb_sheet.cell(row=row_idx, column=date_columns[closest_begin_date] + 1)
                    end_cell = tb_sheet.cell(row=row_idx, column=date_columns[closest_end_date] + 1)
                    
                    # Get Financial Statement Classification using the identified column
                    fin_statement_class = tb_sheet.cell(row=row_idx, column=fin_statement_col).value
                    
                    
                    # Handle different types appropriately
                    if fin_statement_class is None:
                        fin_statement_class = ''
                    else:
                        fin_statement_class = str(fin_statement_class).strip()
                    
                    data.append({
                        'Account Id': display_account_id,  # For display in output
                        'Original Account Id': original_account_id,  # For TB-DATA lookup
                        'Account Name': str(account_name).strip() if account_name is not None else '',
                        'Beginning Balance': 0.0,  # Will be populated from TB-DATA in trial balance creation
                        'Ending Balance': 0.0,     # Will be populated from TB-DATA in trial balance creation
                        'Financial Statement Classification': fin_statement_class
                    })
                except Exception as e:
                    continue
            
            # Convert to DataFrame
            tb_data = pd.DataFrame(data)
            self.print_and_log(f"✅ Extracted {len(data)} account records from TB sheet")
            
            return tb_data
            
        except Exception as e:
            self.print_and_log(f"Error in _extract_trial_balance_data: {str(e)}")
            raise
        finally:
            # Close the workbook if it exists
            try:
                if 'wb' in locals():
                    wb.close()
                    self.print_and_log("Workbook closed")
            except:
                pass

    def _find_financial_classification_column(self, tb_sheet):
        """Find the Financial Statement Classification column in the TB sheet"""
        self.print_and_log("Searching for Financial Statement Classification column...")
        fin_statement_col = None
        
        # Search more efficiently - check fewer rows and columns first
        for row_idx in range(1, 6):  # Check first 5 rows for headers
            self.print_and_log(f"Checking row {row_idx} for headers...")
            for col_idx in range(1, 15):  # Check first 14 columns (reduced from 25)
                try:
                    cell_value = tb_sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        cell_text = str(cell_value).lower()
                        if 'financial statement classification' in cell_text or 'fin statement classification' in cell_text:
                            fin_statement_col = col_idx
                            self.print_and_log(f"Found Financial Statement Classification header in row {row_idx}, column {col_idx}: {cell_value}")
                            return fin_statement_col
                except Exception as e:
                    # Silently continue if there's an error reading a cell
                    continue
            if fin_statement_col:
                break
        
        if not fin_statement_col:
            # Default to column C (3) as specified
            fin_statement_col = 3
            self.print_and_log(f"No Financial Statement Classification header found, using default column C (3)")
        
        return fin_statement_col

    def _create_trial_balance_sheet(self, workbook, trial_balance, styles):
        """Create and format the Comparative Trial Balances sheet"""
        # Create trial balance sheet
        tb_sheet = workbook.create_sheet('Comparative Trial Balances')
        
        # Set column widths for trial balance (converting px to Excel units)
        tb_sheet.column_dimensions['A'].width = 14.3  # 100px
        tb_sheet.column_dimensions['B'].width = 34.3  # 240px
        tb_sheet.column_dimensions['C'].width = 15.7  # 110px
        tb_sheet.column_dimensions['D'].width = 15.7  # 110px
        tb_sheet.column_dimensions['E'].width = 15.7  # 110px
        tb_sheet.column_dimensions['F'].width = 34.3  # 240px
        tb_sheet.column_dimensions['G'].width = 34.3  # 240px
        
        # Write headers
        tb_sheet.append(['Required'] * 5 + ['Optional'] * 2)
        tb_sheet.append(list(trial_balance.columns))
        
        # Style row 1 headers (Required/Optional)
        for col in range(1, 6):  # Columns A-E
            cell = tb_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['blue_fill']
            cell.alignment = styles['center_alignment']
        
        for col in range(6, 8):  # Columns F-G
            cell = tb_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['gray_fill']
            cell.alignment = styles['center_alignment']
        
        # Style row 2 headers (column names) and set row height
        tb_sheet.row_dimensions[2].height = 25  # 33px ≈ 25 points
        for col in range(1, 8):  # All columns
            cell = tb_sheet.cell(row=2, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['dark_blue_fill']
            cell.alignment = styles['center_alignment']
        
        # Write trial balance data with error handling
        for _, row in trial_balance.iterrows():
            row_values = []
            for col in trial_balance.columns:
                value = row[col]
                try:
                    # Double-check the value is clean
                    if isinstance(value, str) and len(value) > 1000:
                        value = value[:1000]
                    row_values.append(value)
                except:
                    row_values.append("ERROR")
            tb_sheet.append(row_values)
        
        return tb_sheet

    def _create_journal_entries_sheets(self, workbook, journal_entries_dict, styles):
        """Create and format multiple Journal Entries & Lines sheets, one for each TXN sheet"""
        # Create journal entries sheets
        created_sheets = []
        
        # Sort the sheets by name to ensure consistent ordering
        sheet_names = sorted(journal_entries_dict.keys())
        
        for idx, sheet_name in enumerate(sheet_names, 1):
            journal_entries = journal_entries_dict[sheet_name]
            sheet_title = f'Journal Entries & Lines {idx}'
            
            # Create the sheet
            je_sheet = workbook.create_sheet(sheet_title)
            
            # Set column widths for journal entries (converting px to Excel units)
            je_sheet.column_dimensions['A'].width = 14.3  # 100px
            je_sheet.column_dimensions['B'].width = 48.6  # 340px
            je_sheet.column_dimensions['C'].width = 15.7  # 110px
            je_sheet.column_dimensions['D'].width = 20.0  # 140px
            je_sheet.column_dimensions['E'].width = 35.7  # 250px
            je_sheet.column_dimensions['F'].width = 15.7  # 110px
            je_sheet.column_dimensions['G'].width = 15.7  # 110px
            je_sheet.column_dimensions['H'].width = 15.7  # 110px
            je_sheet.column_dimensions['I'].width = 15.7  # 110px
            
            # Write headers
            je_sheet.append(['Required', 'Optional', 'Optional', 'Required', 'Required', 'Required', 'Optional', 'Required', 'Required'])
            je_sheet.append(['Journal ID', 'Type', 'Journal Entry Description', 'Posted Date', 'Account ID', 'Journal Line Description', 'Name', 'Debit Amount', 'Credit Amount'])
            
            # Style row 1 headers (Required/Optional)
            required_cols = [1, 4, 5, 6, 8, 9]  # Columns A, D, E, F, H, I
            optional_cols = [2, 3, 7]  # Columns B, C, G
            
            for col in required_cols:
                cell = je_sheet.cell(row=1, column=col)
                cell.font = styles['header_font']
                cell.fill = styles['blue_fill']
                cell.alignment = styles['center_alignment']
            
            for col in optional_cols:
                cell = je_sheet.cell(row=1, column=col)
                cell.font = styles['header_font']
                cell.fill = styles['gray_fill']
                cell.alignment = styles['center_alignment']
            
            # Style row 2 headers (column names)
            for col in range(1, 10):  # All columns
                cell = je_sheet.cell(row=2, column=col)
                cell.font = styles['header_font']
                cell.fill = styles['dark_blue_fill']
                cell.alignment = styles['center_alignment']
            
            # Write journal entries data with error handling
            row_idx = 3  # Start after headers
            chunk_size = 1000  # Process in chunks to avoid memory issues
            total_rows = len(journal_entries)
            chunks_processed = 0
            
            # Process the DataFrame in chunks
            for chunk_start in range(0, total_rows, chunk_size):
                chunk_end = min(chunk_start + chunk_size, total_rows)
                chunk = journal_entries.iloc[chunk_start:chunk_end]
                
                for _, row in chunk.iterrows():
                    try:
                        # Convert row values to appropriate types
                        values = [
                            str(row['Journal ID']),
                            str(row['Type']),
                            str(row['Journal Entry Description']),
                            row['Posted Date'],  # Keep as datetime for formatting
                            str(row['Account ID']),
                            str(row['Journal Line Description']),
                            str(row['Name']),
                            float(row['Debit Amount']),
                            float(row['Credit Amount'])
                        ]
                        
                        # Write values to cells
                        for col_idx, value in enumerate(values, 1):
                            cell = je_sheet.cell(row=row_idx, column=col_idx)
                            cell.value = value
                            
                            # Apply date formatting to Posted Date column
                            if col_idx == 4:  # Posted Date column
                                cell.number_format = 'M/D/YYYY'
                            
                            # Apply number formatting to Debit/Credit columns
                            elif col_idx in [8, 9]:  # Debit/Credit columns
                                cell.number_format = '#,##0.00'
                        
                        row_idx += 1
                        
                    except Exception as e:
                        self.print_and_log(f"Error writing row {row_idx}: {str(e)}")
                        continue
                
                chunks_processed += 1
                if chunks_processed % 10 == 0:  # Log progress every 10 chunks
                    self.print_and_log(f"  • Processed {chunk_end} of {total_rows} rows in {sheet_title}")
            
            created_sheets.append(je_sheet)
            self.print_and_log(f"Created sheet '{sheet_title}' from {sheet_name} with {row_idx-2} transactions")
        
        return created_sheets

    def _create_non_usd_transactions_sheet(self, workbook, styles):
        """Create Non-USD Transactions sheet with Journal Entries columns plus Transaction Currency"""
        non_usd_sheet = workbook.create_sheet('Non-USD Transactions')
        
        if not self.non_usd_transactions:
            return non_usd_sheet
        
        # Set column widths (optimized for Journal Entries format)
        non_usd_sheet.column_dimensions['A'].width = 14.3  # Journal ID
        non_usd_sheet.column_dimensions['B'].width = 48.6  # Type
        non_usd_sheet.column_dimensions['C'].width = 15.7  # Journal Entry Description
        non_usd_sheet.column_dimensions['D'].width = 20.0  # Posted Date
        non_usd_sheet.column_dimensions['E'].width = 35.7  # Account ID
        non_usd_sheet.column_dimensions['F'].width = 15.7  # Journal Line Description
        non_usd_sheet.column_dimensions['G'].width = 15.7  # Name
        non_usd_sheet.column_dimensions['H'].width = 15.7  # Debit Amount
        non_usd_sheet.column_dimensions['I'].width = 15.7  # Credit Amount
        non_usd_sheet.column_dimensions['J'].width = 15.7  # Transaction Currency
        
        # Write headers (Journal Entries & Lines columns + Transaction Currency)
        non_usd_sheet.append(['Required', 'Optional', 'Optional', 'Required', 'Required', 'Required', 'Optional', 'Required', 'Required', 'Required'])
        non_usd_sheet.append(self.non_usd_headers)
        
        # Style row 1 headers (Required/Optional)
        required_cols = [1, 4, 5, 6, 8, 9, 10]  # Columns A, D, E, F, H, I, J
        optional_cols = [2, 3, 7]  # Columns B, C, G
        
        for col in required_cols:
            cell = non_usd_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['blue_fill']
            cell.alignment = styles['center_alignment']
        
        for col in optional_cols:
            cell = non_usd_sheet.cell(row=1, column=col)
            cell.font = styles['header_font']
            cell.fill = styles['gray_fill']
            cell.alignment = styles['center_alignment']
        
        # Style row 2 headers (column names)
        for col_idx in range(1, len(self.non_usd_headers) + 1):
            cell = non_usd_sheet.cell(row=2, column=col_idx)
            cell.font = styles['header_font']
            cell.fill = styles['dark_blue_fill']
            cell.alignment = styles['center_alignment']
        
        # Write non-USD transaction data
        for transaction in self.non_usd_transactions:
            row_data = []
            for header in self.non_usd_headers:
                value = transaction.get(header, '')
                # Clean the value for Excel
                if isinstance(value, str) and len(value) > 1000:
                    value = value[:1000]
                row_data.append(value)
            non_usd_sheet.append(row_data)
        
        # Apply date formatting to Posted Date column (column D)
        for row_num in range(3, non_usd_sheet.max_row + 1):
            cell = non_usd_sheet.cell(row=row_num, column=4)
            cell.number_format = 'M/D/YYYY'
        
        self.print_and_log(f"Created Non-USD Transactions sheet with {len(self.non_usd_transactions)} transactions")
        return non_usd_sheet

    def _extract_balances_from_tb_data(self, account_id, begin_date, end_date):
        """Extract beginning and ending balances from TB-DATA sheet for a specific account"""
        if 'TB-DATA' not in self.source_data or self.source_data['TB-DATA'] is None:
            return 0.0, 0.0

        tb_data = self.source_data['TB-DATA']

        # Find the account in TB-DATA
        account_data = None

        # Check columns B through D (indices 1-3) for Account ID
        for col_idx in [1, 2, 3]:
            try:
                if col_idx < len(tb_data.columns):
                    matching_rows = tb_data[tb_data.iloc[:, col_idx].astype(str) == str(account_id)]
                    if not matching_rows.empty:
                        account_data = matching_rows
                        break
            except Exception:
                continue

        if account_data is None or account_data.empty:
            return 0.0, 0.0

        # Get all available fiscal months for this account and sort them
        fiscal_months = []
        balances_by_month = {}
        
        for _, row in account_data.iterrows():
            try:
                fiscal_month_val = row.iloc[0]
                if pd.isna(fiscal_month_val):
                    continue
                fiscal_month = pd.to_datetime(fiscal_month_val)
                starting_balance = row.iloc[4] if len(row) > 4 else 0.0
                ending_balance = row.iloc[6] if len(row) > 6 else 0.0
                try:
                    starting_balance = float(starting_balance) if pd.notna(starting_balance) else 0.0
                except (ValueError, TypeError):
                    starting_balance = 0.0
                try:
                    ending_balance = float(ending_balance) if pd.notna(ending_balance) else 0.0
                except (ValueError, TypeError):
                    ending_balance = 0.0
                fiscal_months.append(fiscal_month)
                balances_by_month[fiscal_month] = {
                    'start': starting_balance,
                    'end': ending_balance
                }
            except Exception:
                continue

        if not fiscal_months:
            return 0.0, 0.0

        # Sort the fiscal months
        fiscal_months.sort()

        # Always use the earliest for beginning, latest for ending
        begin_fiscal_month = fiscal_months[0]
        end_fiscal_month = fiscal_months[-1]
        begin_balance = balances_by_month[begin_fiscal_month]['start']
        end_balance = balances_by_month[end_fiscal_month]['end']
        return begin_balance, end_balance

if __name__ == "__main__":
    parser = StrongboxParser()
    parser.run() 